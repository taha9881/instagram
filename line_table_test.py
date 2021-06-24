#import all necessary library's 
from pandas.core.frame import DataFrame
import dash
import dash_html_components as html
import dash_bootstrap_components as dbc
import dash_core_components as dcc
import plotly.express as px
import pandas as pd
#
import os
import pathlib
import numpy as np
from datetime import datetime as jj
import multiprocessing
import time
import datetime as dt 
from decimal import Decimal
from openpyxl import load_workbook
import yfinance as yf
from tabulate import tabulate
import xlsxwriter
import plotly.figure_factory as ff
import matplotlib.pyplot as plt
import seaborn as sns
import dash_table




#runProgram function will run for every timeframe at the same time
def runProgram(timeframe_value,path_to_write,time_to_wait,loop):
    
    
    refresh=0
    
    while refresh <=loop:
        start_time=time.time()
        today_name = dt.date.today().strftime("%Y-%m-%d")

        df_cs_analysis = pd.DataFrame([])
        global old_data,new_data
        if refresh == 0 :         
            old_data=pd.DataFrame()
            new_data=pd.DataFrame()   
            old_data= get_value(timeframe_value)
            new_data =get_value(timeframe_value)
            
        
            
          



        ######## Indicator VARIABLES ########

        # RSI
        RSI_VALUE = 14

        # MACD
        MACD_a = 12
        MACD_b = 26
        MACD_c = 9

        # Donchian Channel
        DC_Period = 20

        # Trendfilter Moving Average
        MA_Slow = 200

        ######## USER VARIABLES ########

        # definition of user input
        timeframe_dict = {
            '5m':2, # max 6 days
            '15m':4, # max 6 days
            '30m':7, # max 6 days
            '60m':13, # max 6 days
            '1d':300, # max 6 days
            '1wk':2100, # max 6 days
            '1mo':8400, # max 6 days
        }

        for key, value in timeframe_dict.items():
            TIMEFRAME = key
            LOOKBACK_PERIOD = value
            if (TIMEFRAME ==timeframe_value):

            # TIMEFRAME = '30m' #6 days maximum lookback
            # LOOKBACK_PERIOD = 6 #max 6 days

                START = dt.datetime.today() - dt.timedelta(days=LOOKBACK_PERIOD)
                END = dt.datetime.today()

                TIMESTR = time.strftime("%Y%m%d_%H%M")

                ######## OPTIONS ########

                # all different symbols (pairs) to check
                pair_dict = {
                    'AUD/CAD':5,
                    'AUD/CHF':5,
                    'AUD/JPY':3,
                    'AUD/NZD':5,
                    'AUD/USD':5,
                    'CAD/CHF':3,
                    'CAD/JPY':3,
                    'CHF/JPY':3,
                    'EUR/AUD':5,
                    'EUR/CAD':5,
                    'EUR/CHF':5,
                    'EUR/GBP':5,
                    'EUR/JPY':3,
                    'EUR/NZD':5,
                    'EUR/USD':5,
                    'GBP/AUD':5,
                    'GBP/CAD':5,
                    'GBP/CHF':5,
                    'GBP/JPY':3,
                    'GBP/NZD':5,
                    'GBP/USD':5,
                    'NZD/CAD':5,
                    'NZD/CHF':5,
                    'NZD/JPY':3,
                    'NZD/USD':5,
                    'USD/CAD':5,
                    'USD/CHF':5,
                    'USD/JPY':3,
                }

                ######## FUNCTIONS ########

                # function for extracting data
                def get_data(ticker, start, end, ival):
                    ohlcv_data = {}

                    ohlcv_data[ticker] = yf.download(ticker, start, end, interval=ival)

                    ohlcv_data[ticker]['Ticker'] = ticker

                    df_ticker = ohlcv_data[ticker].copy()

                    df_ticker['date_date'] = [d.date() for d in df_ticker.index]
                    df_ticker['date_time'] = [d.time() for d in df_ticker.index]
                    df_ticker['datetime'] = df_ticker.index
                    df_ticker['datetime'] = df_ticker['datetime'].apply(lambda x: dt.datetime.strftime(x, '%Y-%m-%d %H:%M:%S'))

                    return df_ticker


                def rsi(df, n):
                    "function to calculate RSI"
                    delta = df["Adj Close"].diff().dropna()
                    u = delta * 0
                    d = u.copy()
                    u[delta > 0] = delta[delta > 0]
                    d[delta < 0] = -delta[delta < 0]
                    u[u.index[n-1]] = np.mean( u[:n]) # first value is average of gains
                    u = u.drop(u.index[:(n-1)])
                    d[d.index[n-1]] = np.mean( d[:n]) # first value is average of losses
                    d = d.drop(d.index[:(n-1)])
                    rs = u.ewm(com=n,min_periods=n).mean()/d.ewm(com=n,min_periods=n).mean()
                    return 100 - 100 / (1+rs)

                def RSI(DF,n):
                    "function to calculate RSI"
                    df = DF.copy()
                    df['delta']=df['Adj Close'] - df['Adj Close'].shift(1)
                    df['gain']=np.where(df['delta']>=0,df['delta'],0)
                    df['loss']=np.where(df['delta']<0,abs(df['delta']),0)
                    avg_gain = []
                    avg_loss = []
                    gain = df['gain'].tolist()
                    loss = df['loss'].tolist()
                    for i in range(len(df)):
                        if i < n:
                            avg_gain.append(np.NaN)
                            avg_loss.append(np.NaN)
                        elif i == n:
                            avg_gain.append(df['gain'].rolling(n).mean().tolist()[n])
                            avg_loss.append(df['loss'].rolling(n).mean().tolist()[n])
                        elif i > n:
                            avg_gain.append(((n-1)*avg_gain[i-1] + gain[i])/n)
                            avg_loss.append(((n-1)*avg_loss[i-1] + loss[i])/n)
                    df['avg_gain']=np.array(avg_gain)
                    df['avg_loss']=np.array(avg_loss)
                    df['RS'] = df['avg_gain']/df['avg_loss']
                    df['RSI'] = 100 - (100/(1+df['RS']))
                    return df['RSI']



                def MACD(DF,a,b,c):
                    """function to calculate MACD
                    typical values a = 12; b =26, c =9"""
                    df = DF.copy()
                    df["MA_Fast"]=df["Adj Close"].ewm(span=a,min_periods=a).mean()
                    df["MA_Slow"]=df["Adj Close"].ewm(span=b,min_periods=b).mean()
                    df["MACD"]=df["MA_Fast"]-df["MA_Slow"]
                    df["Signal"]=df["MACD"].ewm(span=c,min_periods=c).mean()
                    df.dropna(inplace=True)
                    return df


                ######## START EXTRACTING HISTORICAL DATA ########

                # initialize result df
                eval_df = pd.DataFrame([])

                # for SYMBOL in pair_list:
                for SYMBOL in pair_dict:

                    print('\n######################')
                    print('Starting with Symbol: ' + SYMBOL)

                    DECIMAL = pair_dict.get(SYMBOL)

                    print('... Decimals: ' + str(DECIMAL))

                    # initialize temporary df
                    df_to_xls = pd.DataFrame([])
                    df_to_xls_inv = pd.DataFrame([])

                    PAIR = SYMBOL.replace('/', '') + '=X'

                    # get historical data from yahoo api
                    data = get_data(PAIR, START, END, TIMEFRAME)

                    data['timeframe'] = TIMEFRAME
                    data['symbol'] = SYMBOL
                    data['Open'] = round(data['Open'],DECIMAL)
                    data['High'] = round(data['High'],DECIMAL)
                    data['Low'] = round(data['Low'],DECIMAL)
                    data['Close'] = round(data['Close'],DECIMAL)
                    data['Adj Close'] = round(data['Adj Close'],DECIMAL)

                    data['RSI'] = RSI(data,14)
                    data["MA_Fast"]=data["Adj Close"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                    data["MA_Slow"]=data["Adj Close"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                    data["MACD"]=data["MA_Fast"]-data["MA_Slow"]
                    data["Signal"]=data["MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean()
                    data["MACD_Diff"]=data["MACD"]-data["Signal"]
                    data["DC_Low"] = data["Low"].rolling(window=DC_Period).min()
                    data["DC_High"] = data["High"].rolling(window=DC_Period).max()
                    data["DC_Base"] = data["DC_Low"] + ((data["DC_High"] - data["DC_High"])/2)
                    data["EMA_200"] = data['Adj Close'].ewm(span=MA_Slow).mean()
                    data["SMA_200"] = data['Adj Close'].rolling(MA_Slow).mean()

                    df_to_xls = data.copy()
                    df_to_xls_inv = data.copy()

                    df_to_xls_inv['timeframe'] = TIMEFRAME
                    df_to_xls_inv['symbol'] = SYMBOL[-3:]+'/'+SYMBOL[:3]
                    df_to_xls_inv['Open'] = round(1/data['Open'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['High'] = round(1/data['High'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['Low'] = round(1/data['Low'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['Close'] = round(1/data['Close'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['Adj Close'] = round(1/data['Adj Close'],DECIMAL if DECIMAL == 5 else 6)

                #     df_to_xls_inv['RSI'] = rsi(df_to_xls_inv,RSI_VALUE)
                    df_to_xls_inv['RSI'] = RSI(df_to_xls_inv,RSI_VALUE)
                    df_to_xls_inv["MA_Fast"]=df_to_xls_inv["Adj Close"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                    df_to_xls_inv["MA_Slow"]=df_to_xls_inv["Adj Close"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                    df_to_xls_inv["MACD"]=df_to_xls_inv["MA_Fast"]-df_to_xls_inv["MA_Slow"]
                    df_to_xls_inv["Signal"]=df_to_xls_inv["MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean()
                    df_to_xls_inv["MACD_Diff"]=df_to_xls_inv["MACD"]-df_to_xls_inv["Signal"]
                    df_to_xls_inv["DC_Low"] = df_to_xls_inv["Low"].rolling(window=DC_Period).min()
                    df_to_xls_inv["DC_High"] = df_to_xls_inv["High"].rolling(window=DC_Period).max()
                    df_to_xls_inv["DC_Base"] = df_to_xls_inv["DC_Low"] + ((df_to_xls_inv["DC_High"] - df_to_xls_inv["DC_High"])/2)
                    df_to_xls_inv["EMA_200"] = df_to_xls_inv['Adj Close'].ewm(span=MA_Slow).mean()
                    df_to_xls_inv["SMA_200"] = df_to_xls_inv['Adj Close'].rolling(MA_Slow).mean()
                    if timeframe_value in ['30m','1d']:
                        eval_df = eval_df.append(df_to_xls[['timeframe',
                                                        'datetime',
                                                        'date_date',
                                                        'date_time',
                                                        'symbol',
                                                        'Open',
                                                        'High',
                                                        'Low',
                                                        'Close',
                                                        'Adj Close',
                                                        'RSI',
                                                        'MACD',
                                                        'Signal',
                                                        'MACD_Diff',
                                                        'DC_Low',
                                                        'DC_High',
                                                        'DC_Base',
                                                        'EMA_200',
                                                        'SMA_200',
                                                        ]], ignore_index=True)

                        eval_df = eval_df.append(df_to_xls_inv[['timeframe',
                                                            'datetime',
                                                            'date_date',
                                                            'date_time',
                                                            'symbol',
                                                            'Open',
                                                            'High',
                                                            'Low',
                                                            'Close',
                                                            'Adj Close',
                                                            'RSI',
                                                            'MACD',
                                                            'Signal',
                                                            'MACD_Diff',
                                                            'DC_Low',
                                                            'DC_High',
                                                            'DC_Base',
                                                            'EMA_200',
                                                            'SMA_200',
                                                        ]], ignore_index=True)
                    else:
                        eval_df = eval_df.append(df_to_xls[['timeframe',
                                                        'datetime',
                                                        'date_date',
                                                        'date_time',
                                                        'symbol',
                                                        'Open',
                                                        'High',
                                                        'Low',
                                                        'Close',
                                                        'Adj Close',
                                                        'RSI',
                                                        'MACD',
                                                        'Signal',
                                                        'MACD_Diff',
                                                        'DC_Low',
                                                        'DC_High',
                                                        'DC_Base',
                                                        'EMA_200',
                                                        'SMA_200',
                                                        ]][:-1], ignore_index=True)

                        eval_df = eval_df.append(df_to_xls_inv[['timeframe',
                                                            'datetime',
                                                            'date_date',
                                                            'date_time',
                                                            'symbol',
                                                            'Open',
                                                            'High',
                                                            'Low',
                                                            'Close',
                                                            'Adj Close',
                                                            'RSI',
                                                            'MACD',
                                                            'Signal',
                                                            'MACD_Diff',
                                                            'DC_Low',
                                                            'DC_High',
                                                            'DC_Base',
                                                            'EMA_200',
                                                            'SMA_200',
                                                        ]][:-1], ignore_index=True)




                time.sleep(1)

                symbol_list = ["EUR", "USD", "GBP", "AUD", "NZD", "JPY", "CAD", "CHF"]
                col_names = ["strength_eur", "strength_usd", "strength_gbp", "strength_aud", "strength_nzd", "strength_jpy", "strength_cad", "strength_chf"]
                center_value = 50
                result_rsi = pd.DataFrame()


            # RSI based Currency strength calculation

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'RSI':'mean'})
                    df[col] = round(df['RSI'] - center_value,2)
                    df = df.drop('RSI', 1)
                    result_rsi = result_rsi.join(df, how='outer')
                    result_rsi['method'] = 'rsi'


                result_rsi['strength_eur_impulse'] = result_rsi['strength_eur']-result_rsi['strength_eur'].shift(1)
                result_rsi['strength_usd_impulse'] = result_rsi['strength_usd']-result_rsi['strength_usd'].shift(1)
                result_rsi['strength_gbp_impulse'] = result_rsi['strength_gbp']-result_rsi['strength_gbp'].shift(1)
                result_rsi['strength_aud_impulse'] = result_rsi['strength_aud']-result_rsi['strength_aud'].shift(1)
                result_rsi['strength_nzd_impulse'] = result_rsi['strength_nzd']-result_rsi['strength_nzd'].shift(1)
                result_rsi['strength_jpy_impulse'] = result_rsi['strength_jpy']-result_rsi['strength_jpy'].shift(1)
                result_rsi['strength_cad_impulse'] = result_rsi['strength_cad']-result_rsi['strength_cad'].shift(1)
                result_rsi['strength_chf_impulse'] = result_rsi['strength_chf']-result_rsi['strength_chf'].shift(1)


            # MACD based Currency strength calculation

                result_macd = pd.DataFrame()

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'MACD':'mean'})
                    df[col] = round(df['MACD'] *10000,2)
                    df = df.drop('MACD', 1)
                    result_macd = result_macd.join(df, how='outer')
                    result_macd['method'] = 'macd'

                result_macd_signal = pd.DataFrame()

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'Signal':'mean'})
                    df[col] = round(df['Signal'] *10000,2)
                    df = df.drop('Signal', 1)
                    result_macd_signal = result_macd_signal.join(df, how='outer')
                    result_macd_signal['method'] = 'macd_signal'

                result_macd_diff = pd.DataFrame()

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'MACD_Diff':'mean'})
                    df[col] = round(df['MACD_Diff'] *10000,2)
                    df = df.drop('MACD_Diff', 1)
                    result_macd_diff = result_macd_diff.join(df, how='outer')
                    result_macd_diff['method'] = 'macd_diff'

                result_macd['strength_eur_impulse'] = result_macd['strength_eur']-result_macd['strength_eur'].shift(1)
                result_macd['strength_usd_impulse'] = result_macd['strength_usd']-result_macd['strength_usd'].shift(1)
                result_macd['strength_gbp_impulse'] = result_macd['strength_gbp']-result_macd['strength_gbp'].shift(1)
                result_macd['strength_aud_impulse'] = result_macd['strength_aud']-result_macd['strength_aud'].shift(1)
                result_macd['strength_nzd_impulse'] = result_macd['strength_nzd']-result_macd['strength_nzd'].shift(1)
                result_macd['strength_jpy_impulse'] = result_macd['strength_jpy']-result_macd['strength_jpy'].shift(1)
                result_macd['strength_cad_impulse'] = result_macd['strength_cad']-result_macd['strength_cad'].shift(1)
                result_macd['strength_chf_impulse'] = result_macd['strength_chf']-result_macd['strength_chf'].shift(1)

                result_macd_signal['strength_eur_impulse'] = result_macd_signal['strength_eur']-result_macd_signal['strength_eur'].shift(1)
                result_macd_signal['strength_usd_impulse'] = result_macd_signal['strength_usd']-result_macd_signal['strength_usd'].shift(1)
                result_macd_signal['strength_gbp_impulse'] = result_macd_signal['strength_gbp']-result_macd_signal['strength_gbp'].shift(1)
                result_macd_signal['strength_aud_impulse'] = result_macd_signal['strength_aud']-result_macd_signal['strength_aud'].shift(1)
                result_macd_signal['strength_nzd_impulse'] = result_macd_signal['strength_nzd']-result_macd_signal['strength_nzd'].shift(1)
                result_macd_signal['strength_jpy_impulse'] = result_macd_signal['strength_jpy']-result_macd_signal['strength_jpy'].shift(1)
                result_macd_signal['strength_cad_impulse'] = result_macd_signal['strength_cad']-result_macd_signal['strength_cad'].shift(1)
                result_macd_signal['strength_chf_impulse'] = result_macd_signal['strength_chf']-result_macd_signal['strength_chf'].shift(1)

                result_macd_diff['strength_eur_impulse'] = result_macd_diff['strength_eur']-result_macd_diff['strength_eur'].shift(1)
                result_macd_diff['strength_usd_impulse'] = result_macd_diff['strength_usd']-result_macd_diff['strength_usd'].shift(1)
                result_macd_diff['strength_gbp_impulse'] = result_macd_diff['strength_gbp']-result_macd_diff['strength_gbp'].shift(1)
                result_macd_diff['strength_aud_impulse'] = result_macd_diff['strength_aud']-result_macd_diff['strength_aud'].shift(1)
                result_macd_diff['strength_nzd_impulse'] = result_macd_diff['strength_nzd']-result_macd_diff['strength_nzd'].shift(1)
                result_macd_diff['strength_jpy_impulse'] = result_macd_diff['strength_jpy']-result_macd_diff['strength_jpy'].shift(1)
                result_macd_diff['strength_cad_impulse'] = result_macd_diff['strength_cad']-result_macd_diff['strength_cad'].shift(1)
                result_macd_diff['strength_chf_impulse'] = result_macd_diff['strength_chf']-result_macd_diff['strength_chf'].shift(1)


            # Putting Milestones together

                df_new_result = pd.DataFrame([])

                df_new_result = df_new_result.append(result_rsi.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)
                df_new_result = df_new_result.append(result_macd.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)
                df_new_result = df_new_result.append(result_macd_signal.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)
                df_new_result = df_new_result.append(result_macd_diff.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)

                ema_length_fast = 9
                ema_length_slow = 21

                col_fast_names = ["strength_eur_ema_fast", "strength_usd_ema_fast", "strength_gbp_ema_fast", "strength_aud_ema_fast", "strength_nzd_ema_fast",
                                "strength_jpy_ema_fast", "strength_cad_ema_fast", "strength_chf_ema_fast"]
                col_slow_names = ["strength_eur_ema_slow", "strength_usd_ema_slow", "strength_gbp_ema_slow", "strength_aud_ema_slow", "strength_nzd_ema_slow",
                                "strength_jpy_ema_slow", "strength_cad_ema_slow", "strength_chf_ema_slow"]

                for col, fast, slow in zip(col_names, col_fast_names, col_slow_names):
                    result_rsi[fast] = result_rsi[col].rolling(window=ema_length_fast, min_periods=1).mean()
                    result_rsi[slow] = result_rsi[col].rolling(window=ema_length_slow, min_periods=1).mean()
                    result_rsi[fast] = round(result_rsi[fast],2)
                    result_rsi[slow] = round(result_rsi[slow],2)

                result_rsi.sort_values(by=['datetime'], ascending=False)

                result_rsi['strength_eur_ema_fast_impulse'] = result_rsi['strength_eur_ema_fast']-result_rsi['strength_eur_ema_fast'].shift(1)
                result_rsi['strength_usd_ema_fast_impulse'] = result_rsi['strength_usd_ema_fast']-result_rsi['strength_usd_ema_fast'].shift(1)
                result_rsi['strength_gbp_ema_fast_impulse'] = result_rsi['strength_gbp_ema_fast']-result_rsi['strength_gbp_ema_fast'].shift(1)
                result_rsi['strength_aud_ema_fast_impulse'] = result_rsi['strength_aud_ema_fast']-result_rsi['strength_aud_ema_fast'].shift(1)
                result_rsi['strength_nzd_ema_fast_impulse'] = result_rsi['strength_nzd_ema_fast']-result_rsi['strength_nzd_ema_fast'].shift(1)
                result_rsi['strength_jpy_ema_fast_impulse'] = result_rsi['strength_jpy_ema_fast']-result_rsi['strength_jpy_ema_fast'].shift(1)
                result_rsi['strength_cad_ema_fast_impulse'] = result_rsi['strength_cad_ema_fast']-result_rsi['strength_cad_ema_fast'].shift(1)
                result_rsi['strength_chf_ema_fast_impulse'] = result_rsi['strength_chf_ema_fast']-result_rsi['strength_chf_ema_fast'].shift(1)

                result_rsi['strength_eur_ema_slow_impulse'] = result_rsi['strength_eur_ema_slow']-result_rsi['strength_eur_ema_slow'].shift(1)
                result_rsi['strength_usd_ema_slow_impulse'] = result_rsi['strength_usd_ema_slow']-result_rsi['strength_usd_ema_slow'].shift(1)
                result_rsi['strength_gbp_ema_slow_impulse'] = result_rsi['strength_gbp_ema_slow']-result_rsi['strength_gbp_ema_slow'].shift(1)
                result_rsi['strength_aud_ema_slow_impulse'] = result_rsi['strength_aud_ema_slow']-result_rsi['strength_aud_ema_slow'].shift(1)
                result_rsi['strength_nzd_ema_slow_impulse'] = result_rsi['strength_nzd_ema_slow']-result_rsi['strength_nzd_ema_slow'].shift(1)
                result_rsi['strength_jpy_ema_slow_impulse'] = result_rsi['strength_jpy_ema_slow']-result_rsi['strength_jpy_ema_slow'].shift(1)
                result_rsi['strength_cad_ema_slow_impulse'] = result_rsi['strength_cad_ema_slow']-result_rsi['strength_cad_ema_slow'].shift(1)
                result_rsi['strength_chf_ema_slow_impulse'] = result_rsi['strength_chf_ema_slow']-result_rsi['strength_chf_ema_slow'].shift(1)

                result_rsi["strength_eur_MA_Fast"] = result_rsi["strength_eur"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_eur_MA_Slow"] = result_rsi["strength_eur"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_eur_MACD"] = round(result_rsi["strength_eur_MA_Fast"]-result_rsi["strength_eur_MA_Slow"],3)
                result_rsi["strength_eur_Signal"] = round(result_rsi["strength_eur_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_eur_MACD_Diff"] = round(result_rsi["strength_eur_MACD"]-result_rsi["strength_eur_Signal"],3)
                result_rsi['strength_eur_MACD_Diff_impulse'] = round(result_rsi['strength_eur_MACD_Diff']-result_rsi['strength_eur_MACD_Diff'].shift(1),3)

                result_rsi["strength_usd_MA_Fast"] = result_rsi["strength_usd"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_usd_MA_Slow"] = result_rsi["strength_usd"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_usd_MACD"] = round(result_rsi["strength_usd_MA_Fast"]-result_rsi["strength_usd_MA_Slow"],3)
                result_rsi["strength_usd_Signal"] = round(result_rsi["strength_usd_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_usd_MACD_Diff"] = round(result_rsi["strength_usd_MACD"]-result_rsi["strength_usd_Signal"],3)
                result_rsi['strength_usd_MACD_Diff_impulse'] = round(result_rsi['strength_usd_MACD_Diff']-result_rsi['strength_usd_MACD_Diff'].shift(1),3)

                result_rsi["strength_gbp_MA_Fast"] = result_rsi["strength_gbp"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_gbp_MA_Slow"] = result_rsi["strength_gbp"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_gbp_MACD"] = round(result_rsi["strength_gbp_MA_Fast"]-result_rsi["strength_gbp_MA_Slow"],3)
                result_rsi["strength_gbp_Signal"] = round(result_rsi["strength_gbp_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_gbp_MACD_Diff"] = round(result_rsi["strength_gbp_MACD"]-result_rsi["strength_gbp_Signal"],3)
                result_rsi['strength_gbp_MACD_Diff_impulse'] = round(result_rsi['strength_gbp_MACD_Diff']-result_rsi['strength_gbp_MACD_Diff'].shift(1),3)

                result_rsi["strength_aud_MA_Fast"] = result_rsi["strength_aud"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_aud_MA_Slow"] = result_rsi["strength_aud"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_aud_MACD"] = round(result_rsi["strength_aud_MA_Fast"]-result_rsi["strength_aud_MA_Slow"],3)
                result_rsi["strength_aud_Signal"] = round(result_rsi["strength_aud_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_aud_MACD_Diff"] = round(result_rsi["strength_aud_MACD"]-result_rsi["strength_aud_Signal"],3)
                result_rsi['strength_aud_MACD_Diff_impulse'] = round(result_rsi['strength_aud_MACD_Diff']-result_rsi['strength_aud_MACD_Diff'].shift(1),3)

                result_rsi["strength_nzd_MA_Fast"] = result_rsi["strength_nzd"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_nzd_MA_Slow"] = result_rsi["strength_nzd"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_nzd_MACD"] = round(result_rsi["strength_nzd_MA_Fast"]-result_rsi["strength_nzd_MA_Slow"],3)
                result_rsi["strength_nzd_Signal"] = round(result_rsi["strength_nzd_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_nzd_MACD_Diff"] = round(result_rsi["strength_nzd_MACD"]-result_rsi["strength_nzd_Signal"],3)
                result_rsi['strength_nzd_MACD_Diff_impulse'] = round(result_rsi['strength_nzd_MACD_Diff']-result_rsi['strength_nzd_MACD_Diff'].shift(1),3)

                result_rsi["strength_jpy_MA_Fast"] = result_rsi["strength_jpy"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_jpy_MA_Slow"] = result_rsi["strength_jpy"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_jpy_MACD"] = round(result_rsi["strength_jpy_MA_Fast"]-result_rsi["strength_jpy_MA_Slow"],3)
                result_rsi["strength_jpy_Signal"] = round(result_rsi["strength_jpy_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_jpy_MACD_Diff"] = round(result_rsi["strength_jpy_MACD"]-result_rsi["strength_jpy_Signal"],3)
                result_rsi['strength_jpy_MACD_Diff_impulse'] = round(result_rsi['strength_jpy_MACD_Diff']-result_rsi['strength_jpy_MACD_Diff'].shift(1),3)

                result_rsi["strength_cad_MA_Fast"] = result_rsi["strength_cad"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_cad_MA_Slow"] = result_rsi["strength_cad"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_cad_MACD"] = round(result_rsi["strength_cad_MA_Fast"]-result_rsi["strength_cad_MA_Slow"],3)
                result_rsi["strength_cad_Signal"] = round(result_rsi["strength_cad_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_cad_MACD_Diff"] = round(result_rsi["strength_cad_MACD"]-result_rsi["strength_cad_Signal"],3)
                result_rsi['strength_cad_MACD_Diff_impulse'] = round(result_rsi['strength_cad_MACD_Diff']-result_rsi['strength_cad_MACD_Diff'].shift(1),3)

                result_rsi["strength_chf_MA_Fast"] = result_rsi["strength_chf"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_chf_MA_Slow"] = result_rsi["strength_chf"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_chf_MACD"] = round(result_rsi["strength_chf_MA_Fast"]-result_rsi["strength_chf_MA_Slow"],3)
                result_rsi["strength_chf_Signal"] = round(result_rsi["strength_chf_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_chf_MACD_Diff"] = round(result_rsi["strength_chf_MACD"]-result_rsi["strength_chf_Signal"],3)
                result_rsi['strength_chf_MACD_Diff_impulse'] = round(result_rsi['strength_chf_MACD_Diff']-result_rsi['strength_chf_MACD_Diff'].shift(1),3)

                ### RSI Strength Calculation
                df_strength = pd.DataFrame([])

                # create new dataframe for "safe" calculations
                df_strength = result_rsi.copy()

                # drop empty values
                df_strength.dropna()

                # sort dataframe descending to get last datetime / most recent timestamp and transpose df
                df_strength_transposed = df_strength.sort_values(by=['datetime'], ascending=False)[:1].T.copy()

                # create new numeric index to "change" old index into "normal column"
                df_strength_transposed.reset_index(inplace=True)
                df_strength_transposed['datetime'] = df_strength_transposed.columns[1]

                # rename columns
                df_strength_transposed.rename(columns={df_strength_transposed.columns[0]: "Currency",
                                                    df_strength_transposed.columns[1]: "Strength",
                                                    }, inplace=True)

                # rename index column in most recent timestamp VALUE
                # (i can´t implement it and i´m not even sure if it is a good idea - but we could change it back if it´s not)
                df_strength_transposed.rename_axis("Timestamp", axis="columns", inplace=True)

                # sort dataframe by currency
                df_strength_transposed.sort_values(by=['Currency'], ascending=True, inplace=True)

                # reset index to match sorted order
                df_strength_transposed.reset_index(drop=True, inplace=True)

                df_strength_transposed['Currency'] = df_strength_transposed['Currency'].str.replace('strength_', '')

                df_strength_transposed['category'] = [x[:3] for x in df_strength_transposed['Currency']]
                df_strength_transposed['header'] = [x[4:] for x in df_strength_transposed['Currency']]

                df_strength = pd.DataFrame([])

                df_strength = df_strength_transposed[(df_strength_transposed['Currency']!='method')].copy()

                df_strength_aud = df_strength[(df_strength['category']=='aud')][['header','Currency','Strength']].T.copy()
                df_strength_aud.columns = df_strength_aud.iloc[0]
                df_strength_aud.reset_index(drop=True, inplace=True)
                df_strength_aud.rename_axis("Index", axis="columns", inplace=True)
                df_strength_aud.rename(columns={df_strength_aud.columns[0]: "strength",}, inplace=True)
                df_strength_aud['currency'] = 'aud'

                df_strength_cad = df_strength[(df_strength['category']=='cad')][['header','Currency','Strength']].T.copy()
                df_strength_cad.columns = df_strength_cad.iloc[0]
                df_strength_cad.reset_index(drop=True, inplace=True)
                df_strength_cad.rename_axis("Index", axis="columns", inplace=True)
                df_strength_cad.rename(columns={df_strength_cad.columns[0]: "strength",}, inplace=True)
                df_strength_cad['currency'] = 'cad'

                df_strength_chf = df_strength[(df_strength['category']=='chf')][['header','Currency','Strength']].T.copy()
                df_strength_chf.columns = df_strength_chf.iloc[0]
                df_strength_chf.reset_index(drop=True, inplace=True)
                df_strength_chf.rename_axis("Index", axis="columns", inplace=True)
                df_strength_chf.rename(columns={df_strength_chf.columns[0]: "strength",}, inplace=True)
                df_strength_chf['currency'] = 'chf'

                df_strength_eur = df_strength[(df_strength['category']=='eur')][['header','Currency','Strength']].T.copy()
                df_strength_eur.columns = df_strength_eur.iloc[0]
                df_strength_eur.reset_index(drop=True, inplace=True)
                df_strength_eur.rename_axis("Index", axis="columns", inplace=True)
                df_strength_eur.rename(columns={df_strength_eur.columns[0]: "strength",}, inplace=True)
                df_strength_eur['currency'] = 'eur'

                df_strength_gbp = df_strength[(df_strength['category']=='gbp')][['header','Currency','Strength']].T.copy()
                df_strength_gbp.columns = df_strength_gbp.iloc[0]
                df_strength_gbp.reset_index(drop=True, inplace=True)
                df_strength_gbp.rename_axis("Index", axis="columns", inplace=True)
                df_strength_gbp.rename(columns={df_strength_gbp.columns[0]: "strength",}, inplace=True)
                df_strength_gbp['currency'] = 'gbp'

                df_strength_nzd = df_strength[(df_strength['category']=='nzd')][['header','Currency','Strength']].T.copy()
                df_strength_nzd.columns = df_strength_nzd.iloc[0]
                df_strength_nzd.reset_index(drop=True, inplace=True)
                df_strength_nzd.rename_axis("Index", axis="columns", inplace=True)
                df_strength_nzd.rename(columns={df_strength_nzd.columns[0]: "strength",}, inplace=True)
                df_strength_nzd['currency'] = 'nzd'

                df_strength_jpy = df_strength[(df_strength['category']=='jpy')][['header','Currency','Strength']].T.copy()
                df_strength_jpy.columns = df_strength_jpy.iloc[0]
                df_strength_jpy.reset_index(drop=True, inplace=True)
                df_strength_jpy.rename_axis("Index", axis="columns", inplace=True)
                df_strength_jpy.rename(columns={df_strength_jpy.columns[0]: "strength",}, inplace=True)
                df_strength_jpy['currency'] = 'jpy'

                df_strength_usd = df_strength[(df_strength['category']=='usd')][['header','Currency','Strength']].T.copy()
                df_strength_usd.columns = df_strength_usd.iloc[0]
                df_strength_usd.reset_index(drop=True, inplace=True)
                df_strength_usd.rename_axis("Index", axis="columns", inplace=True)
                df_strength_usd.rename(columns={df_strength_usd.columns[0]: "strength",}, inplace=True)
                df_strength_usd['currency'] = 'usd'

                df_strength_result = pd.DataFrame([])

                df_strength_result = df_strength_result.append(df_strength_aud[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_cad[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_chf[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_eur[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_gbp[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_nzd[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_jpy[2:3], ignore_index=True)
                df_strength_result = df_strength_result.append(df_strength_usd[2:3], ignore_index=True)

                df_strength_result['datetime'] = df_strength_transposed['datetime']
                df_strength_result['timeframe'] = TIMEFRAME

                df_strength_result.sort_values(by=['strength'], ascending=False, inplace=True)
                df_strength_result.reset_index(drop=True, inplace=True)

                df_strength_export = df_strength_result[['timeframe',
                                                        'datetime',
                                                        'currency',
                                                        'impulse',
                                                        'strength',
                                                        'ema_fast',
                                                        'ema_slow',
                                                        'ema_fast_impulse',
                                                        'ema_slow_impulse',
                                                        'MACD',
                                                        'Signal',
                                                        'MACD_Diff',
                                                        'MACD_Diff_impulse',
                                                        ]].copy()

                # final result
                df_cs_analysis = df_cs_analysis.append(df_strength_export, ignore_index=True)
                print("Interval "+ TIMEFRAME + " finished.")

        # Save result to Excel document
        print("\n\nSaving results to XLSX")
        out_path = path_to_write + '_CS.xlsx'

        #Append new record or Create new excel file for perticular timeframe
        def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                            truncate_sheet=False, 
                            **to_excel_kwargs):
            
            
            # Excel file doesn't exist - saving and exiting
            if not os.path.isfile(filename):
                df.to_excel(
                    filename,
                    sheet_name=sheet_name, 
                    startrow=startrow if startrow is not None else 0, 
                    **to_excel_kwargs)
                return
            
            # ignore [engine] parameter if it was passed
            if 'engine' in to_excel_kwargs:
                to_excel_kwargs.pop('engine')
            
            

            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

            # try to open an existing workbook
            writer.book = load_workbook(filename)
            
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            
            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

            if startrow is None:
                startrow = 0

            # write out the new sheet
            df.to_excel(writer, sheet_name, startrow=startrow,header=None, **to_excel_kwargs)

            # save the workbook
            writer.save()
        
        #calling appending function
        append_df_to_excel(out_path,df_cs_analysis,"CS Report")
        
        #Get on sleep as per the timeframe
        refresh+=1
        if timeframe_value =='5m':
            df=get_value('5m')

            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["5m strength"]= new_data['5m strength']
            new_data['5m strength'] = df["5m strength"]

            old_csv_data["5m strength"]=old_data["5m strength"]
            new_csv_data['5m strength']=new_data["5m strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)
            

            
        elif timeframe_value =='15m':
            df=get_value('15m')            
            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["15m strength"]= new_data['15m strength']
            new_data['15m strength'] = df["15m strength"]

            old_csv_data["15m strength"]=old_data["15m strength"]
            new_csv_data['15m strength']=new_data["15m strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)

        elif timeframe_value =='30m':
            df=get_value('30m')                        
            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["30m strength"]= new_data['30m strength']
            new_data['30m strength'] = df["30m strength"]

            old_csv_data["30m strength"]=old_data["30m strength"]
            new_csv_data['30m strength']=new_data["30m strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)

            
        
        elif timeframe_value =='60m':
            df=get_value('60m')            
            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["60m strength"]= new_data['60m strength']
            new_data['60m strength'] = df["60m strength"]

            old_csv_data["60m strength"]=old_data["60m strength"]
            new_csv_data['60m strength']=new_data["60m strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)


        elif timeframe_value =='1d':
            df=get_value('1d')            
            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["1d strength"]= new_data['1d strength']
            new_data['1d strength'] = df["1d strength"]

            old_csv_data["1d strength"]=old_data["1d strength"]
            new_csv_data['1d strength']=new_data["1d strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)
          
        elif timeframe_value =='1wk':
            df=get_value('1wk')            
            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["1wk strength"]= new_data['1wk strength']
            new_data['1wk strength'] = df["1wk strength"]

            old_csv_data["1wk strength"]=old_data["1wk strength"]
            new_csv_data['1wk strength']=new_data["1wk strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)
        elif timeframe_value =='1mo':
            df=get_value('1mo')            
            old_csv_data = pd.read_csv('old_data.csv')
            new_csv_data = pd.read_csv('new_data.csv')      

            old_data["1mo strength"]= new_data['1mo strength']
            new_data['1mo strength'] = df["1mo strength"]

            old_csv_data["1mo strength"]=old_data["1mo strength"]
            new_csv_data['1mo strength']=new_data["1mo strength"]

            old_csv_data.to_csv("old_data.csv",index=False)
            new_csv_data.to_csv("new_data.csv",index=False)
        

        end_time=time.time()
        time_delta=end_time - start_time
        temp_time_to_wait = time_to_wait - time_delta
        print("Sleeping for {} sec".format(temp_time_to_wait))
        time.sleep(temp_time_to_wait)

#Histtorical data generator for each timeframe

def historical_data_generator(timeframe_value,path_to_write,loop,n_previous_instance):
    

    # date as string for filenames
    refresh=0
    while  refresh <=loop:
    
        
        today_name = dt.date.today().strftime("%Y-%m-%d")

        df_cs_analysis = pd.DataFrame([])


        ######## Indicator VARIABLES ########

        # RSI
        RSI_VALUE = 14

        # MACD
        MACD_a = 12
        MACD_b = 26
        MACD_c = 9

        # Donchian Channel
        DC_Period = 20

        # Trendfilter Moving Average
        MA_Slow = 200

        ######## USER VARIABLES ########

        # definition of user input
        timeframe_dict = {
            '5m':2, # max 6 days
            '15m':4, # max 6 days
            '30m':7, # max 6 days
            '60m':13, # max 6 days
            '1d':300, # max 6 days
            '1wk':2100, # max 6 days
            '1mo':8400, # max 6 days
        }

        for key, value in timeframe_dict.items():
            TIMEFRAME = key
            LOOKBACK_PERIOD = value
            if (TIMEFRAME ==timeframe_value):

            # TIMEFRAME = '30m' #6 days maximum lookback
            # LOOKBACK_PERIOD = 6 #max 6 days

                START = dt.datetime.today() - dt.timedelta(days=LOOKBACK_PERIOD)
                END = dt.datetime.today()

                TIMESTR = time.strftime("%Y%m%d_%H%M")

                ######## OPTIONS ########

                # all different symbols (pairs) to check
                pair_dict = {
                    'AUD/CAD':5,
                    'AUD/CHF':5,
                    'AUD/JPY':3,
                    'AUD/NZD':5,
                    'AUD/USD':5,
                    'CAD/CHF':3,
                    'CAD/JPY':3,
                    'CHF/JPY':3,
                    'EUR/AUD':5,
                    'EUR/CAD':5,
                    'EUR/CHF':5,
                    'EUR/GBP':5,
                    'EUR/JPY':3,
                    'EUR/NZD':5,
                    'EUR/USD':5,
                    'GBP/AUD':5,
                    'GBP/CAD':5,
                    'GBP/CHF':5,
                    'GBP/JPY':3,
                    'GBP/NZD':5,
                    'GBP/USD':5,
                    'NZD/CAD':5,
                    'NZD/CHF':5,
                    'NZD/JPY':3,
                    'NZD/USD':5,
                    'USD/CAD':5,
                    'USD/CHF':5,
                    'USD/JPY':3,
                }

                ######## FUNCTIONS ########

                # function for extracting data
                def get_data(ticker, start, end, ival):
                    ohlcv_data = {}

                    ohlcv_data[ticker] = yf.download(ticker, start, end, interval=ival)

                    ohlcv_data[ticker]['Ticker'] = ticker

                    df_ticker = ohlcv_data[ticker].copy()

                    df_ticker['date_date'] = [d.date() for d in df_ticker.index]
                    df_ticker['date_time'] = [d.time() for d in df_ticker.index]
                    df_ticker['datetime'] = df_ticker.index
                    df_ticker['datetime'] = df_ticker['datetime'].apply(lambda x: dt.datetime.strftime(x, '%Y-%m-%d %H:%M:%S'))

                    return df_ticker


                def rsi(df, n):
                    "function to calculate RSI"
                    delta = df["Adj Close"].diff().dropna()
                    u = delta * 0
                    d = u.copy()
                    u[delta > 0] = delta[delta > 0]
                    d[delta < 0] = -delta[delta < 0]
                    u[u.index[n-1]] = np.mean( u[:n]) # first value is average of gains
                    u = u.drop(u.index[:(n-1)])
                    d[d.index[n-1]] = np.mean( d[:n]) # first value is average of losses
                    d = d.drop(d.index[:(n-1)])
                    rs = u.ewm(com=n,min_periods=n).mean()/d.ewm(com=n,min_periods=n).mean()
                    return 100 - 100 / (1+rs)

                def RSI(DF,n):
                    "function to calculate RSI"
                    df = DF.copy()
                    df['delta']=df['Adj Close'] - df['Adj Close'].shift(1)
                    df['gain']=np.where(df['delta']>=0,df['delta'],0)
                    df['loss']=np.where(df['delta']<0,abs(df['delta']),0)
                    avg_gain = []
                    avg_loss = []
                    gain = df['gain'].tolist()
                    loss = df['loss'].tolist()
                    for i in range(len(df)):
                        if i < n:
                            avg_gain.append(np.NaN)
                            avg_loss.append(np.NaN)
                        elif i == n:
                            avg_gain.append(df['gain'].rolling(n).mean().tolist()[n])
                            avg_loss.append(df['loss'].rolling(n).mean().tolist()[n])
                        elif i > n:
                            avg_gain.append(((n-1)*avg_gain[i-1] + gain[i])/n)
                            avg_loss.append(((n-1)*avg_loss[i-1] + loss[i])/n)
                    df['avg_gain']=np.array(avg_gain)
                    df['avg_loss']=np.array(avg_loss)
                    df['RS'] = df['avg_gain']/df['avg_loss']
                    df['RSI'] = 100 - (100/(1+df['RS']))
                    return df['RSI']



                def MACD(DF,a,b,c):
                    """function to calculate MACD
                    typical values a = 12; b =26, c =9"""
                    df = DF.copy()
                    df["MA_Fast"]=df["Adj Close"].ewm(span=a,min_periods=a).mean()
                    df["MA_Slow"]=df["Adj Close"].ewm(span=b,min_periods=b).mean()
                    df["MACD"]=df["MA_Fast"]-df["MA_Slow"]
                    df["Signal"]=df["MACD"].ewm(span=c,min_periods=c).mean()
                    df.dropna(inplace=True)
                    return df


                ######## START EXTRACTING HISTORICAL DATA ########

                # initialize result df
                eval_df = pd.DataFrame([])

                # for SYMBOL in pair_list:
                for SYMBOL in pair_dict:

                    # print('\n######################')
                    # print('Starting with Symbol: ' + SYMBOL)

                    DECIMAL = pair_dict.get(SYMBOL)

                    # print('... Decimals: ' + str(DECIMAL))

                    # initialize temporary df
                    df_to_xls = pd.DataFrame([])
                    df_to_xls_inv = pd.DataFrame([])

                    PAIR = SYMBOL.replace('/', '') + '=X'

                    # get historical data from yahoo api
                    data = get_data(PAIR, START, END, TIMEFRAME)

                    data['timeframe'] = TIMEFRAME
                    data['symbol'] = SYMBOL
                    data['Open'] = round(data['Open'],DECIMAL)
                    data['High'] = round(data['High'],DECIMAL)
                    data['Low'] = round(data['Low'],DECIMAL)
                    data['Close'] = round(data['Close'],DECIMAL)
                    data['Adj Close'] = round(data['Adj Close'],DECIMAL)

                    data['RSI'] = RSI(data,14)
                    data["MA_Fast"]=data["Adj Close"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                    data["MA_Slow"]=data["Adj Close"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                    data["MACD"]=data["MA_Fast"]-data["MA_Slow"]
                    data["Signal"]=data["MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean()
                    data["MACD_Diff"]=data["MACD"]-data["Signal"]
                    data["DC_Low"] = data["Low"].rolling(window=DC_Period).min()
                    data["DC_High"] = data["High"].rolling(window=DC_Period).max()
                    data["DC_Base"] = data["DC_Low"] + ((data["DC_High"] - data["DC_High"])/2)
                    data["EMA_200"] = data['Adj Close'].ewm(span=MA_Slow).mean()
                    data["SMA_200"] = data['Adj Close'].rolling(MA_Slow).mean()

                    df_to_xls = data.copy()
                    df_to_xls_inv = data.copy()

                    df_to_xls_inv['timeframe'] = TIMEFRAME
                    df_to_xls_inv['symbol'] = SYMBOL[-3:]+'/'+SYMBOL[:3]
                    df_to_xls_inv['Open'] = round(1/data['Open'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['High'] = round(1/data['High'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['Low'] = round(1/data['Low'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['Close'] = round(1/data['Close'],DECIMAL if DECIMAL == 5 else 6)
                    df_to_xls_inv['Adj Close'] = round(1/data['Adj Close'],DECIMAL if DECIMAL == 5 else 6)

                #     df_to_xls_inv['RSI'] = rsi(df_to_xls_inv,RSI_VALUE)
                    df_to_xls_inv['RSI'] = RSI(df_to_xls_inv,RSI_VALUE)
                    df_to_xls_inv["MA_Fast"]=df_to_xls_inv["Adj Close"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                    df_to_xls_inv["MA_Slow"]=df_to_xls_inv["Adj Close"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                    df_to_xls_inv["MACD"]=df_to_xls_inv["MA_Fast"]-df_to_xls_inv["MA_Slow"]
                    df_to_xls_inv["Signal"]=df_to_xls_inv["MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean()
                    df_to_xls_inv["MACD_Diff"]=df_to_xls_inv["MACD"]-df_to_xls_inv["Signal"]
                    df_to_xls_inv["DC_Low"] = df_to_xls_inv["Low"].rolling(window=DC_Period).min()
                    df_to_xls_inv["DC_High"] = df_to_xls_inv["High"].rolling(window=DC_Period).max()
                    df_to_xls_inv["DC_Base"] = df_to_xls_inv["DC_Low"] + ((df_to_xls_inv["DC_High"] - df_to_xls_inv["DC_High"])/2)
                    df_to_xls_inv["EMA_200"] = df_to_xls_inv['Adj Close'].ewm(span=MA_Slow).mean()
                    df_to_xls_inv["SMA_200"] = df_to_xls_inv['Adj Close'].rolling(MA_Slow).mean()

                    if timeframe_value in ['30m','1d']:
                        eval_df = eval_df.append(df_to_xls[['timeframe',
                                                        'datetime',
                                                        'date_date',
                                                        'date_time',
                                                        'symbol',
                                                        'Open',
                                                        'High',
                                                        'Low',
                                                        'Close',
                                                        'Adj Close',
                                                        'RSI',
                                                        'MACD',
                                                        'Signal',
                                                        'MACD_Diff',
                                                        'DC_Low',
                                                        'DC_High',
                                                        'DC_Base',
                                                        'EMA_200',
                                                        'SMA_200',
                                                        ]], ignore_index=True)

                        eval_df = eval_df.append(df_to_xls_inv[['timeframe',
                                                            'datetime',
                                                            'date_date',
                                                            'date_time',
                                                            'symbol',
                                                            'Open',
                                                            'High',
                                                            'Low',
                                                            'Close',
                                                            'Adj Close',
                                                            'RSI',
                                                            'MACD',
                                                            'Signal',
                                                            'MACD_Diff',
                                                            'DC_Low',
                                                            'DC_High',
                                                            'DC_Base',
                                                            'EMA_200',
                                                            'SMA_200',
                                                        ]], ignore_index=True)
                    else:
                        eval_df = eval_df.append(df_to_xls[['timeframe',
                                                        'datetime',
                                                        'date_date',
                                                        'date_time',
                                                        'symbol',
                                                        'Open',
                                                        'High',
                                                        'Low',
                                                        'Close',
                                                        'Adj Close',
                                                        'RSI',
                                                        'MACD',
                                                        'Signal',
                                                        'MACD_Diff',
                                                        'DC_Low',
                                                        'DC_High',
                                                        'DC_Base',
                                                        'EMA_200',
                                                        'SMA_200',
                                                        ]][:-1], ignore_index=True)

                        eval_df = eval_df.append(df_to_xls_inv[['timeframe',
                                                            'datetime',
                                                            'date_date',
                                                            'date_time',
                                                            'symbol',
                                                            'Open',
                                                            'High',
                                                            'Low',
                                                            'Close',
                                                            'Adj Close',
                                                            'RSI',
                                                            'MACD',
                                                            'Signal',
                                                            'MACD_Diff',
                                                            'DC_Low',
                                                            'DC_High',
                                                            'DC_Base',
                                                            'EMA_200',
                                                            'SMA_200',
                                                        ]][:-1], ignore_index=True)



                time.sleep(1)

                symbol_list = ["EUR", "USD", "GBP", "AUD", "NZD", "JPY", "CAD", "CHF"]
                col_names = ["strength_eur", "strength_usd", "strength_gbp", "strength_aud", "strength_nzd", "strength_jpy", "strength_cad", "strength_chf"]
                center_value = 50
                result_rsi = pd.DataFrame()


            # RSI based Currency strength calculation

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'RSI':'mean'})
                    df[col] = round(df['RSI'] - center_value,2)
                    df = df.drop('RSI', 1)
                    result_rsi = result_rsi.join(df, how='outer')
                    result_rsi['method'] = 'rsi'


                result_rsi['strength_eur_impulse'] = result_rsi['strength_eur']-result_rsi['strength_eur'].shift(1)
                result_rsi['strength_usd_impulse'] = result_rsi['strength_usd']-result_rsi['strength_usd'].shift(1)
                result_rsi['strength_gbp_impulse'] = result_rsi['strength_gbp']-result_rsi['strength_gbp'].shift(1)
                result_rsi['strength_aud_impulse'] = result_rsi['strength_aud']-result_rsi['strength_aud'].shift(1)
                result_rsi['strength_nzd_impulse'] = result_rsi['strength_nzd']-result_rsi['strength_nzd'].shift(1)
                result_rsi['strength_jpy_impulse'] = result_rsi['strength_jpy']-result_rsi['strength_jpy'].shift(1)
                result_rsi['strength_cad_impulse'] = result_rsi['strength_cad']-result_rsi['strength_cad'].shift(1)
                result_rsi['strength_chf_impulse'] = result_rsi['strength_chf']-result_rsi['strength_chf'].shift(1)


            # MACD based Currency strength calculation

                result_macd = pd.DataFrame()

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'MACD':'mean'})
                    df[col] = round(df['MACD'] *10000,2)
                    df = df.drop('MACD', 1)
                    result_macd = result_macd.join(df, how='outer')
                    result_macd['method'] = 'macd'

                result_macd_signal = pd.DataFrame()

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'Signal':'mean'})
                    df[col] = round(df['Signal'] *10000,2)
                    df = df.drop('Signal', 1)
                    result_macd_signal = result_macd_signal.join(df, how='outer')
                    result_macd_signal['method'] = 'macd_signal'

                result_macd_diff = pd.DataFrame()

                for s, col in zip(symbol_list, col_names):
                    df = eval_df[eval_df['symbol'].apply(lambda x: x.startswith(s))]
                    df = df.groupby('datetime').agg({'MACD_Diff':'mean'})
                    df[col] = round(df['MACD_Diff'] *10000,2)
                    df = df.drop('MACD_Diff', 1)
                    result_macd_diff = result_macd_diff.join(df, how='outer')
                    result_macd_diff['method'] = 'macd_diff'

                result_macd['strength_eur_impulse'] = result_macd['strength_eur']-result_macd['strength_eur'].shift(1)
                result_macd['strength_usd_impulse'] = result_macd['strength_usd']-result_macd['strength_usd'].shift(1)
                result_macd['strength_gbp_impulse'] = result_macd['strength_gbp']-result_macd['strength_gbp'].shift(1)
                result_macd['strength_aud_impulse'] = result_macd['strength_aud']-result_macd['strength_aud'].shift(1)
                result_macd['strength_nzd_impulse'] = result_macd['strength_nzd']-result_macd['strength_nzd'].shift(1)
                result_macd['strength_jpy_impulse'] = result_macd['strength_jpy']-result_macd['strength_jpy'].shift(1)
                result_macd['strength_cad_impulse'] = result_macd['strength_cad']-result_macd['strength_cad'].shift(1)
                result_macd['strength_chf_impulse'] = result_macd['strength_chf']-result_macd['strength_chf'].shift(1)

                result_macd_signal['strength_eur_impulse'] = result_macd_signal['strength_eur']-result_macd_signal['strength_eur'].shift(1)
                result_macd_signal['strength_usd_impulse'] = result_macd_signal['strength_usd']-result_macd_signal['strength_usd'].shift(1)
                result_macd_signal['strength_gbp_impulse'] = result_macd_signal['strength_gbp']-result_macd_signal['strength_gbp'].shift(1)
                result_macd_signal['strength_aud_impulse'] = result_macd_signal['strength_aud']-result_macd_signal['strength_aud'].shift(1)
                result_macd_signal['strength_nzd_impulse'] = result_macd_signal['strength_nzd']-result_macd_signal['strength_nzd'].shift(1)
                result_macd_signal['strength_jpy_impulse'] = result_macd_signal['strength_jpy']-result_macd_signal['strength_jpy'].shift(1)
                result_macd_signal['strength_cad_impulse'] = result_macd_signal['strength_cad']-result_macd_signal['strength_cad'].shift(1)
                result_macd_signal['strength_chf_impulse'] = result_macd_signal['strength_chf']-result_macd_signal['strength_chf'].shift(1)

                result_macd_diff['strength_eur_impulse'] = result_macd_diff['strength_eur']-result_macd_diff['strength_eur'].shift(1)
                result_macd_diff['strength_usd_impulse'] = result_macd_diff['strength_usd']-result_macd_diff['strength_usd'].shift(1)
                result_macd_diff['strength_gbp_impulse'] = result_macd_diff['strength_gbp']-result_macd_diff['strength_gbp'].shift(1)
                result_macd_diff['strength_aud_impulse'] = result_macd_diff['strength_aud']-result_macd_diff['strength_aud'].shift(1)
                result_macd_diff['strength_nzd_impulse'] = result_macd_diff['strength_nzd']-result_macd_diff['strength_nzd'].shift(1)
                result_macd_diff['strength_jpy_impulse'] = result_macd_diff['strength_jpy']-result_macd_diff['strength_jpy'].shift(1)
                result_macd_diff['strength_cad_impulse'] = result_macd_diff['strength_cad']-result_macd_diff['strength_cad'].shift(1)
                result_macd_diff['strength_chf_impulse'] = result_macd_diff['strength_chf']-result_macd_diff['strength_chf'].shift(1)


            # Putting Milestones together

                df_new_result = pd.DataFrame([])

                df_new_result = df_new_result.append(result_rsi.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)
                df_new_result = df_new_result.append(result_macd.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)
                df_new_result = df_new_result.append(result_macd_signal.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)
                df_new_result = df_new_result.append(result_macd_diff.sort_values(by=['datetime'], ascending=False)[:1], ignore_index=True)

                ema_length_fast = 9
                ema_length_slow = 21

                col_fast_names = ["strength_eur_ema_fast", "strength_usd_ema_fast", "strength_gbp_ema_fast", "strength_aud_ema_fast", "strength_nzd_ema_fast",
                                "strength_jpy_ema_fast", "strength_cad_ema_fast", "strength_chf_ema_fast"]
                col_slow_names = ["strength_eur_ema_slow", "strength_usd_ema_slow", "strength_gbp_ema_slow", "strength_aud_ema_slow", "strength_nzd_ema_slow",
                                "strength_jpy_ema_slow", "strength_cad_ema_slow", "strength_chf_ema_slow"]

                for col, fast, slow in zip(col_names, col_fast_names, col_slow_names):
                    result_rsi[fast] = result_rsi[col].rolling(window=ema_length_fast, min_periods=1).mean()
                    result_rsi[slow] = result_rsi[col].rolling(window=ema_length_slow, min_periods=1).mean()
                    result_rsi[fast] = round(result_rsi[fast],2)
                    result_rsi[slow] = round(result_rsi[slow],2)

                result_rsi.sort_values(by=['datetime'], ascending=False)

                result_rsi['strength_eur_ema_fast_impulse'] = result_rsi['strength_eur_ema_fast']-result_rsi['strength_eur_ema_fast'].shift(1)
                result_rsi['strength_usd_ema_fast_impulse'] = result_rsi['strength_usd_ema_fast']-result_rsi['strength_usd_ema_fast'].shift(1)
                result_rsi['strength_gbp_ema_fast_impulse'] = result_rsi['strength_gbp_ema_fast']-result_rsi['strength_gbp_ema_fast'].shift(1)
                result_rsi['strength_aud_ema_fast_impulse'] = result_rsi['strength_aud_ema_fast']-result_rsi['strength_aud_ema_fast'].shift(1)
                result_rsi['strength_nzd_ema_fast_impulse'] = result_rsi['strength_nzd_ema_fast']-result_rsi['strength_nzd_ema_fast'].shift(1)
                result_rsi['strength_jpy_ema_fast_impulse'] = result_rsi['strength_jpy_ema_fast']-result_rsi['strength_jpy_ema_fast'].shift(1)
                result_rsi['strength_cad_ema_fast_impulse'] = result_rsi['strength_cad_ema_fast']-result_rsi['strength_cad_ema_fast'].shift(1)
                result_rsi['strength_chf_ema_fast_impulse'] = result_rsi['strength_chf_ema_fast']-result_rsi['strength_chf_ema_fast'].shift(1)

                result_rsi['strength_eur_ema_slow_impulse'] = result_rsi['strength_eur_ema_slow']-result_rsi['strength_eur_ema_slow'].shift(1)
                result_rsi['strength_usd_ema_slow_impulse'] = result_rsi['strength_usd_ema_slow']-result_rsi['strength_usd_ema_slow'].shift(1)
                result_rsi['strength_gbp_ema_slow_impulse'] = result_rsi['strength_gbp_ema_slow']-result_rsi['strength_gbp_ema_slow'].shift(1)
                result_rsi['strength_aud_ema_slow_impulse'] = result_rsi['strength_aud_ema_slow']-result_rsi['strength_aud_ema_slow'].shift(1)
                result_rsi['strength_nzd_ema_slow_impulse'] = result_rsi['strength_nzd_ema_slow']-result_rsi['strength_nzd_ema_slow'].shift(1)
                result_rsi['strength_jpy_ema_slow_impulse'] = result_rsi['strength_jpy_ema_slow']-result_rsi['strength_jpy_ema_slow'].shift(1)
                result_rsi['strength_cad_ema_slow_impulse'] = result_rsi['strength_cad_ema_slow']-result_rsi['strength_cad_ema_slow'].shift(1)
                result_rsi['strength_chf_ema_slow_impulse'] = result_rsi['strength_chf_ema_slow']-result_rsi['strength_chf_ema_slow'].shift(1)

                result_rsi["strength_eur_MA_Fast"] = result_rsi["strength_eur"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_eur_MA_Slow"] = result_rsi["strength_eur"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_eur_MACD"] = round(result_rsi["strength_eur_MA_Fast"]-result_rsi["strength_eur_MA_Slow"],3)
                result_rsi["strength_eur_Signal"] = round(result_rsi["strength_eur_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_eur_MACD_Diff"] = round(result_rsi["strength_eur_MACD"]-result_rsi["strength_eur_Signal"],3)
                result_rsi['strength_eur_MACD_Diff_impulse'] = round(result_rsi['strength_eur_MACD_Diff']-result_rsi['strength_eur_MACD_Diff'].shift(1),3)

                result_rsi["strength_usd_MA_Fast"] = result_rsi["strength_usd"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_usd_MA_Slow"] = result_rsi["strength_usd"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_usd_MACD"] = round(result_rsi["strength_usd_MA_Fast"]-result_rsi["strength_usd_MA_Slow"],3)
                result_rsi["strength_usd_Signal"] = round(result_rsi["strength_usd_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_usd_MACD_Diff"] = round(result_rsi["strength_usd_MACD"]-result_rsi["strength_usd_Signal"],3)
                result_rsi['strength_usd_MACD_Diff_impulse'] = round(result_rsi['strength_usd_MACD_Diff']-result_rsi['strength_usd_MACD_Diff'].shift(1),3)

                result_rsi["strength_gbp_MA_Fast"] = result_rsi["strength_gbp"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_gbp_MA_Slow"] = result_rsi["strength_gbp"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_gbp_MACD"] = round(result_rsi["strength_gbp_MA_Fast"]-result_rsi["strength_gbp_MA_Slow"],3)
                result_rsi["strength_gbp_Signal"] = round(result_rsi["strength_gbp_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_gbp_MACD_Diff"] = round(result_rsi["strength_gbp_MACD"]-result_rsi["strength_gbp_Signal"],3)
                result_rsi['strength_gbp_MACD_Diff_impulse'] = round(result_rsi['strength_gbp_MACD_Diff']-result_rsi['strength_gbp_MACD_Diff'].shift(1),3)

                result_rsi["strength_aud_MA_Fast"] = result_rsi["strength_aud"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_aud_MA_Slow"] = result_rsi["strength_aud"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_aud_MACD"] = round(result_rsi["strength_aud_MA_Fast"]-result_rsi["strength_aud_MA_Slow"],3)
                result_rsi["strength_aud_Signal"] = round(result_rsi["strength_aud_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_aud_MACD_Diff"] = round(result_rsi["strength_aud_MACD"]-result_rsi["strength_aud_Signal"],3)
                result_rsi['strength_aud_MACD_Diff_impulse'] = round(result_rsi['strength_aud_MACD_Diff']-result_rsi['strength_aud_MACD_Diff'].shift(1),3)

                result_rsi["strength_nzd_MA_Fast"] = result_rsi["strength_nzd"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_nzd_MA_Slow"] = result_rsi["strength_nzd"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_nzd_MACD"] = round(result_rsi["strength_nzd_MA_Fast"]-result_rsi["strength_nzd_MA_Slow"],3)
                result_rsi["strength_nzd_Signal"] = round(result_rsi["strength_nzd_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_nzd_MACD_Diff"] = round(result_rsi["strength_nzd_MACD"]-result_rsi["strength_nzd_Signal"],3)
                result_rsi['strength_nzd_MACD_Diff_impulse'] = round(result_rsi['strength_nzd_MACD_Diff']-result_rsi['strength_nzd_MACD_Diff'].shift(1),3)

                result_rsi["strength_jpy_MA_Fast"] = result_rsi["strength_jpy"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_jpy_MA_Slow"] = result_rsi["strength_jpy"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_jpy_MACD"] = round(result_rsi["strength_jpy_MA_Fast"]-result_rsi["strength_jpy_MA_Slow"],3)
                result_rsi["strength_jpy_Signal"] = round(result_rsi["strength_jpy_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_jpy_MACD_Diff"] = round(result_rsi["strength_jpy_MACD"]-result_rsi["strength_jpy_Signal"],3)
                result_rsi['strength_jpy_MACD_Diff_impulse'] = round(result_rsi['strength_jpy_MACD_Diff']-result_rsi['strength_jpy_MACD_Diff'].shift(1),3)

                result_rsi["strength_cad_MA_Fast"] = result_rsi["strength_cad"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_cad_MA_Slow"] = result_rsi["strength_cad"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_cad_MACD"] = round(result_rsi["strength_cad_MA_Fast"]-result_rsi["strength_cad_MA_Slow"],3)
                result_rsi["strength_cad_Signal"] = round(result_rsi["strength_cad_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_cad_MACD_Diff"] = round(result_rsi["strength_cad_MACD"]-result_rsi["strength_cad_Signal"],3)
                result_rsi['strength_cad_MACD_Diff_impulse'] = round(result_rsi['strength_cad_MACD_Diff']-result_rsi['strength_cad_MACD_Diff'].shift(1),3)

                result_rsi["strength_chf_MA_Fast"] = result_rsi["strength_chf"].ewm(span=MACD_a,min_periods=MACD_a).mean()
                result_rsi["strength_chf_MA_Slow"] = result_rsi["strength_chf"].ewm(span=MACD_b,min_periods=MACD_b).mean()
                result_rsi["strength_chf_MACD"] = round(result_rsi["strength_chf_MA_Fast"]-result_rsi["strength_chf_MA_Slow"],3)
                result_rsi["strength_chf_Signal"] = round(result_rsi["strength_chf_MACD"].ewm(span=MACD_c,min_periods=MACD_c).mean(),3)
                result_rsi["strength_chf_MACD_Diff"] = round(result_rsi["strength_chf_MACD"]-result_rsi["strength_chf_Signal"],3)
                result_rsi['strength_chf_MACD_Diff_impulse'] = round(result_rsi['strength_chf_MACD_Diff']-result_rsi['strength_chf_MACD_Diff'].shift(1),3)

                ### RSI Strength Calculation
                df_strength = pd.DataFrame([])

                # create new dataframe for "safe" calculations
                df_strength = result_rsi.copy()

                # drop empty values
                df_strength.dropna()

                df_strength=df_strength.sort_values(by=['datetime'], ascending=False)

                # sort dataframe descending to get last datetime / most recent timestamp and transpose df
                df_strength_transposed = df_strength.T.copy()               
                
                while n_previous_instance>=1:
                    

                    temp=df_strength_transposed.iloc[:,[n_previous_instance]]
                    # print(temp)
                    # create new numeric index to "change" old index into "normal column"
                    temp.reset_index(inplace=True)
                    temp['datetime'] = temp.columns[1]

                    # rename columns
                    temp.rename(columns={temp.columns[0]: "Currency",
                                                        temp.columns[1]: "Strength",
                                                        }, inplace=True)

                    # rename index column in most recent timestamp VALUE
                    # (i can´t implement it and i´m not even sure if it is a good idea - but we could change it back if it´s not)
                    temp.rename_axis("Timestamp", axis="columns", inplace=True)

                    # sort dataframe by currency
                    temp.sort_values(by=['Currency'], ascending=True, inplace=True)

                    # reset index to match sorted order
                    temp.reset_index(drop=True, inplace=True)

                    temp['Currency'] = temp['Currency'].str.replace('strength_', '')

                    temp['category'] = [x[:3] for x in temp['Currency']]
                    temp['header'] = [x[4:] for x in temp['Currency']]

                    df_strength = pd.DataFrame([])

                    df_strength = temp[(temp['Currency']!='method')].copy()

                    df_strength_aud = df_strength[(df_strength['category']=='aud')][['header','Currency','Strength']].T.copy()
                    df_strength_aud.columns = df_strength_aud.iloc[0]
                    df_strength_aud.reset_index(drop=True, inplace=True)
                    df_strength_aud.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_aud.rename(columns={df_strength_aud.columns[0]: "strength",}, inplace=True)
                    df_strength_aud['currency'] = 'aud'

                    df_strength_cad = df_strength[(df_strength['category']=='cad')][['header','Currency','Strength']].T.copy()
                    df_strength_cad.columns = df_strength_cad.iloc[0]
                    df_strength_cad.reset_index(drop=True, inplace=True)
                    df_strength_cad.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_cad.rename(columns={df_strength_cad.columns[0]: "strength",}, inplace=True)
                    df_strength_cad['currency'] = 'cad'

                    df_strength_chf = df_strength[(df_strength['category']=='chf')][['header','Currency','Strength']].T.copy()
                    df_strength_chf.columns = df_strength_chf.iloc[0]
                    df_strength_chf.reset_index(drop=True, inplace=True)
                    df_strength_chf.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_chf.rename(columns={df_strength_chf.columns[0]: "strength",}, inplace=True)
                    df_strength_chf['currency'] = 'chf'

                    df_strength_eur = df_strength[(df_strength['category']=='eur')][['header','Currency','Strength']].T.copy()
                    df_strength_eur.columns = df_strength_eur.iloc[0]
                    df_strength_eur.reset_index(drop=True, inplace=True)
                    df_strength_eur.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_eur.rename(columns={df_strength_eur.columns[0]: "strength",}, inplace=True)
                    df_strength_eur['currency'] = 'eur'

                    df_strength_gbp = df_strength[(df_strength['category']=='gbp')][['header','Currency','Strength']].T.copy()
                    df_strength_gbp.columns = df_strength_gbp.iloc[0]
                    df_strength_gbp.reset_index(drop=True, inplace=True)
                    df_strength_gbp.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_gbp.rename(columns={df_strength_gbp.columns[0]: "strength",}, inplace=True)
                    df_strength_gbp['currency'] = 'gbp'

                    df_strength_nzd = df_strength[(df_strength['category']=='nzd')][['header','Currency','Strength']].T.copy()
                    df_strength_nzd.columns = df_strength_nzd.iloc[0]
                    df_strength_nzd.reset_index(drop=True, inplace=True)
                    df_strength_nzd.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_nzd.rename(columns={df_strength_nzd.columns[0]: "strength",}, inplace=True)
                    df_strength_nzd['currency'] = 'nzd'

                    df_strength_jpy = df_strength[(df_strength['category']=='jpy')][['header','Currency','Strength']].T.copy()
                    df_strength_jpy.columns = df_strength_jpy.iloc[0]
                    df_strength_jpy.reset_index(drop=True, inplace=True)
                    df_strength_jpy.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_jpy.rename(columns={df_strength_jpy.columns[0]: "strength",}, inplace=True)
                    df_strength_jpy['currency'] = 'jpy'

                    df_strength_usd = df_strength[(df_strength['category']=='usd')][['header','Currency','Strength']].T.copy()
                    df_strength_usd.columns = df_strength_usd.iloc[0]
                    df_strength_usd.reset_index(drop=True, inplace=True)
                    df_strength_usd.rename_axis("Index", axis="columns", inplace=True)
                    df_strength_usd.rename(columns={df_strength_usd.columns[0]: "strength",}, inplace=True)
                    df_strength_usd['currency'] = 'usd'

                    df_strength_result = pd.DataFrame([])

                    df_strength_result = df_strength_result.append(df_strength_aud[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_cad[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_chf[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_eur[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_gbp[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_nzd[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_jpy[2:3], ignore_index=True)
                    df_strength_result = df_strength_result.append(df_strength_usd[2:3], ignore_index=True)

                    df_strength_result['datetime'] = temp['datetime']
                    df_strength_result['timeframe'] = TIMEFRAME

                    df_strength_result.sort_values(by=['strength'], ascending=False, inplace=True)
                    df_strength_result.reset_index(drop=True, inplace=True)

                    df_strength_export = df_strength_result[['timeframe',
                                                            'datetime',
                                                            'currency',
                                                            'impulse',
                                                            'strength',
                                                            'ema_fast',
                                                            'ema_slow',
                                                            'ema_fast_impulse',
                                                            'ema_slow_impulse',
                                                            'MACD',
                                                            'Signal',
                                                            'MACD_Diff',
                                                            'MACD_Diff_impulse',
                                                            ]].copy()

                    # final result
                    df_cs_analysis = df_cs_analysis.append(df_strength_export, ignore_index=True)
                    # print("Interval "+ TIMEFRAME + " finished.")
                    n_previous_instance-=1

               

        
        out_path = path_to_write + '_CS.xlsx'
        # # Create a Pandas Excel writer using XlsxWriter as the engine.
        writer = pd.ExcelWriter(out_path, engine='xlsxwriter')

        df_cs_analysis.to_excel(writer, sheet_name='CS Report')

        writer.save()
        refresh+=1


#generate table for strength and impulse currency
def show_currency():
    #load excel files
    df1=pd.read_excel("5minfile_CS.xlsx")
    df2=pd.read_excel("15minfile_CS.xlsx")
    df3=pd.read_excel("30minfile_CS.xlsx")
    df4=pd.read_excel("60minfile_CS.xlsx")
    df5=pd.read_excel("1dayfile_CS.xlsx")
    df6=pd.read_excel("1monthfile_CS.xlsx")
    df7=pd.read_excel("1weekfile_CS.xlsx")
    #take last recent values from excel
    x0,x1,x2,x3,x4,x5,x6=df1[-8:],df2[-8:],df3[-8:],df4[-8:],df5[-8:],df6[-8:],df7[-8:]
    #sorting df with currency
    x0.sort_values(by=["currency"],inplace=True)
    x1.sort_values(by=["currency"],inplace=True)
    x2.sort_values(by=["currency"],inplace=True)
    x3.sort_values(by=["currency"],inplace=True)
    x4.sort_values(by=["currency"],inplace=True)
    x5.sort_values(by=["currency"],inplace=True)
    x6.sort_values(by=["currency"],inplace=True)
    #creating df for strength and impulse
    df_s=pd.DataFrame()
    df_i=pd.DataFrame()
    def looping(df,tf,col_num):
        pair=[]
        values=[]
        values2=[]
        #appending values to list
        for i in range(8):
            pair.append(df.iloc[i]['currency'])
            values.append(round(df.iloc[i]["strength"],2))
            values2.append(round(df.iloc[i]["impulse"],2))
        df_i.insert(0,"pair",pair,allow_duplicates=True)
        df_s.insert(0,"pair",pair,allow_duplicates=True)
        df_i.insert(col_num,tf,values2)
        df_s.insert(col_num,tf,values)            
    list2=["5m",'15m','30m','60m','1d','1wk','1mo']
    var_list=[x0,x1,x2,x3,x4,x5,x6]
    col_num=1
    for (x,send) in zip(var_list,list2):
        looping(x,send,col_num)
        col_num+=1
    df_s=df_s.loc[:,~df_s.columns.duplicated()]

    df_i=df_i.loc[:,~df_i.columns.duplicated()]
    df_s=df_s.rename(columns={'1mo':'1mo strength','1wk':'1wk strength','1d':'1d strength','60m':'60m strength','30m':'30m strength','15m':'15m strength','5m':'5m strength'})
    df_i=df_i.loc[:,'1mo':]
    df_i=df_i.rename(columns={'1mo':'1mo impulse','1wk':'1wk impulse','1d':'1d impulse','60m':'60m impulse','30m':'30m impulse','15m':'15m impulse','5m':'5m impulse'})
    #concatinating two dataframe for generating table
    df=pd.concat([df_s,df_i],axis=1)
    return df

#generate table for strength and impulse currency pair values
def show_currency_pair():
    #load excel files
    df1=pd.read_excel("5minfile_CS.xlsx")
    df2=pd.read_excel("15minfile_CS.xlsx")
    df3=pd.read_excel("30minfile_CS.xlsx")
    df4=pd.read_excel("60minfile_CS.xlsx")
    df5=pd.read_excel("1dayfile_CS.xlsx")
    df6=pd.read_excel("1monthfile_CS.xlsx")
    df7=pd.read_excel("1weekfile_CS.xlsx")
    x0,x1,x2,x3,x4,x5,x6=df1[-8:],df2[-8:],df3[-8:],df4[-8:],df5[-8:],df6[-8:],df7[-8:]
    x0.sort_values(by=["currency"],inplace=True)
    x1.sort_values(by=["currency"],inplace=True)
    x2.sort_values(by=["currency"],inplace=True)
    x3.sort_values(by=["currency"],inplace=True)
    x4.sort_values(by=["currency"],inplace=True)
    x5.sort_values(by=["currency"],inplace=True)
    x6.sort_values(by=["currency"],inplace=True)
    #change
    wrong_pair=["jpynzd","jpyusd","chfgbp","chfnzd","chfeur","chfusd","audgbp","cadgbp","audeur","cadnzd","cadeur","cadusd"]
    #creating df for strength and impulse
    df_s=pd.DataFrame()
    df_i=pd.DataFrame()
    def looping(x,timeframe,col_num):    
        pair=[]
        sum_s=[]
        sum_i=[]
        #adding value of pair and appending to list
        for i in range (8):  
            for j in range(i+1,8):
                #change
                if (x.iloc[i]['currency']+x.iloc[j]['currency']) in wrong_pair:
                    pair.append(x.iloc[j]['currency']+x.iloc[i]['currency'])
                    sum_s.append(round((x.iloc[j]['strength'])-(x.iloc[i]['strength']),2))
                    sum_i.append(round((x.iloc[j]['impulse'])-(x.iloc[i]['impulse']),2))
                else:
                    pair.append(x.iloc[i]['currency']+x.iloc[j]['currency']) #change
                    sum_s.append(round((x.iloc[i]['strength'])-(x.iloc[j]['strength']),2))
                    sum_i.append(round((x.iloc[i]['impulse'])-(x.iloc[j]['impulse']),2))
        df_s.insert(0,"pair",pair,allow_duplicates=True)
        df_i.insert(0,"pair",pair,allow_duplicates=True)
        df_s.insert(col_num,timeframe,sum_s)
        df_i.insert(col_num,timeframe,sum_i)
    col_num_send=1
    var_list=[x0,x1,x2,x3,x4,x5,x6]
    list2=["5m",'15m','30m','60m','1d','1wk','1mo']
    for (x,send) in zip(var_list,list2):
        looping(x,send,col_num_send)
        col_num_send+=1
    df_i= df_i.loc[:,~df_i.columns.duplicated()]
    df_i=df_i.loc[:,'1mo':]
    df_i=df_i.rename(columns={'1mo':'1mo impulse','1wk':'1wk impulse','1d':'1d impulse','60m':'60m impulse','30m':'30m impulse','15m':'15m impulse','5m':'5m impulse'})
    df_s=df_s.loc[:,~df_s.columns.duplicated()]
    df_s=df_s.rename(columns={'1mo':'1mo strength','1wk':'1wk strength','1d':'1d strength','60m':'60m strength','30m':'30m strength','15m':'15m strength','5m':'5m strength'})
    #concatinating two dataframe for generating table
    df_pair=pd.concat([df_s,df_i],axis=1)
    return df_pair

#style for Strength values
def discrete_background_color_bins_blues(df, n_bins=5, columns='all'):
    import colorlover
    bounds = [i * (1.0 / n_bins) for i in range(n_bins + 1)]
    if columns == 'all':
        if 'id' in df:
            df_numeric_columns = df.select_dtypes('number').drop(['id'], axis=1)
        else:
            df_numeric_columns = df.select_dtypes('number')
    else:
        df_numeric_columns = df[columns]
    df_max = df_numeric_columns.max().max()
    df_min = df_numeric_columns.min().min()
    ranges = [
        ((df_max - df_min) * i) + df_min
        for i in bounds
    ]
    styles = []
    legend = []
    for i in range(1, len(bounds)):
        min_bound = ranges[i - 1]
        max_bound = ranges[i]
        backgroundColor = colorlover.scales[str(n_bins)]['seq']['Blues'][i - 1]
        color = 'white' if i > len(bounds) / 2. else 'inherit'

        for column in df_numeric_columns:
            styles.append({
                'if': {
                    'filter_query': (
                        '{{{column}}} >= {min_bound}' +
                        (' && {{{column}}} < {max_bound}' if (i < len(bounds) - 1) else '')
                    ).format(column=column, min_bound=min_bound, max_bound=max_bound),
                    'column_id': column
                },
                'backgroundColor': backgroundColor,
                'color': color
            })
        legend.append(
            html.Div(style={'display': 'inline-block', 'width': '60px'}, children=[
                html.Div(
                    style={
                        'backgroundColor': backgroundColor,
                        'borderLeft': '1px rgb(50, 50, 50) solid',
                        'height': '10px'
                    }
                ),
                html.Small(round(min_bound, 2), style={'paddingLeft': '2px'})
            ])
        )

    return (styles, html.Div(legend, style={'padding': '5px 0 5px 0'}))

#style for Impulse values
def discrete_background_color_bins_reds(df, n_bins=5, columns='all'):
    import colorlover
    bounds = [i * (1.0 / n_bins) for i in range(n_bins + 1)]
    if columns == 'all':
        if 'id' in df:
            df_numeric_columns = df.select_dtypes('number').drop(['id'], axis=1)
        else:
            df_numeric_columns = df.select_dtypes('number')
    else:
        df_numeric_columns = df[columns]
    df_max = df_numeric_columns.max().max()
    df_min = df_numeric_columns.min().min()
    ranges = [
        ((df_max - df_min) * i) + df_min
        for i in bounds
    ]
    styles = []
    legend = []
    for i in range(1, len(bounds)):
        min_bound = ranges[i - 1]
        max_bound = ranges[i]
        backgroundColor = colorlover.scales[str(n_bins)]['seq']['Reds'][i - 1]
        color = 'white' if i > len(bounds) / 2. else 'inherit'

        for column in df_numeric_columns:
            styles.append({
                'if': {
                    'filter_query': (
                        '{{{column}}} >= {min_bound}' +
                        (' && {{{column}}} < {max_bound}' if (i < len(bounds) - 1) else '')
                    ).format(column=column, min_bound=min_bound, max_bound=max_bound),
                    'column_id': column
                },
                'backgroundColor': backgroundColor,
                'color': color
            })
        legend.append(
            html.Div(style={'display': 'inline-block', 'width': '60px'}, children=[
                html.Div(
                    style={
                        'backgroundColor': backgroundColor,
                        'borderLeft': '1px rgb(50, 50, 50) solid',
                        'height': '10px'
                    }
                ),
                html.Small(round(min_bound, 2), style={'paddingLeft': '2px'})
            ])
        )

    return (styles, html.Div(legend, style={'padding': '5px 0 5px 0'}))

#get pair value for perticular instance
def get_value(timeframe_main):
    if timeframe_main=='5m':        
        df1=pd.read_excel("5minfile_CS.xlsx")
    elif timeframe_main =='15m':        
        df1=pd.read_excel("15minfile_CS.xlsx")
    elif timeframe_main == '30m':
        df1=pd.read_excel("30minfile_CS.xlsx")
    elif timeframe_main =='60m':
        df1=pd.read_excel("60minfile_CS.xlsx")
    elif timeframe_main == '1d':
        df1=pd.read_excel("1dayfile_CS.xlsx")
    elif timeframe_main =='1mo':
        df1=pd.read_excel("1monthfile_CS.xlsx")
    elif timeframe_main =='1wk':
        df1=pd.read_excel("1weekfile_CS.xlsx")
    
    x1=df1[-8:]
    x1.sort_values(by=["currency"],inplace=True)    
    #change
    wrong_pair=["jpynzd","jpyusd","chfgbp","chfnzd","chfeur","chfusd","audgbp","cadgbp","audeur","cadnzd","cadeur","cadusd"]
    #creating df for strength 
    df_s=pd.DataFrame()
    def looping(x,timeframe,col_num):
        if timeframe == timeframe_main:            
            pair=[]
            sum_s=[]
            #adding value of pair and appending to list
            for i in range (8):  
                for j in range(i+1,8):
                    #change
                    if (x.iloc[i]['currency']+x.iloc[j]['currency']) in wrong_pair:
                        pair.append(x.iloc[j]['currency']+x.iloc[i]['currency'])
                        sum_s.append(round((x.iloc[j]['strength'])-(x.iloc[i]['strength']),2))

                    else:
                        pair.append(x.iloc[i]['currency']+x.iloc[j]['currency']) #change
                        sum_s.append(round((x.iloc[i]['strength'])-(x.iloc[j]['strength']),2))

            df_s.insert(0,"pair",pair,allow_duplicates=True)        
            df_s.insert(col_num,timeframe,sum_s)  
        
    col_num_send=1
    var_list=[x1,x1,x1,x1,x1,x1,x1]
    list2=["5m",'15m','30m','60m','1d','1wk','1mo']
    for (x,send) in zip(var_list,list2):
        looping(x,send,col_num_send)
        col_num_send=1
    df_s=df_s.loc[:,~df_s.columns.duplicated()]
    df_s=df_s.rename(columns={'1mo':'1mo strength','1wk':'1wk strength','1d':'1d strength','60m':'60m strength','30m':'30m strength','15m':'15m strength','5m':'5m strength'})
    return df_s

#get last update data
def last_update():
    df1=pd.read_excel("5minfile_CS.xlsx")
    df2=pd.read_excel("15minfile_CS.xlsx")
    df3=pd.read_excel("30minfile_CS.xlsx")
    df4=pd.read_excel("60minfile_CS.xlsx")
    df5=pd.read_excel("1dayfile_CS.xlsx")
    df6=pd.read_excel("1weekfile_CS.xlsx")
    df7=pd.read_excel("1monthfile_CS.xlsx")
    #take last recent values from excel
    x0,x1,x2,x3,x4,x5,x6=df1[-1:],df2[-1:],df3[-1:],df4[-1:],df5[-1:],df6[-1:],df7[-1:]
    x0=x0[["timeframe","datetime"]]
    x1=x1[["timeframe","datetime"]]
    x2=x2[["timeframe","datetime"]]
    x3=x3[["timeframe","datetime"]]
    x4=x4[["timeframe","datetime"]]
    x5=x5[["timeframe","datetime"]]
    x6=x6[["timeframe","datetime"]]
    new_df= pd.concat([x0,x1,x2,x3,x4,x5,x6])
    datetime=[]
    for dt in new_df["datetime"]:
        temp=dt.split(" ")
        st=temp[1].split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{} {}:{}:{}".format(temp[0],st[0],st[1],st[2])
        datetime.append(temp)
    new_df['datetime']=datetime
    return new_df

#get old update data
def old_update():
    df1=pd.read_excel("5minfile_CS.xlsx")
    df2=pd.read_excel("15minfile_CS.xlsx")
    df3=pd.read_excel("30minfile_CS.xlsx")
    df4=pd.read_excel("60minfile_CS.xlsx")
    df5=pd.read_excel("1dayfile_CS.xlsx")
    df6=pd.read_excel("1weekfile_CS.xlsx")
    df7=pd.read_excel("1monthfile_CS.xlsx")
    #take last recent values from excel
    x0,x1,x2,x3,x4,x5,x6=df1[-1:],df2[-1:],df3[-1:],df4[-1:],df5[-1:],df6[-1:],df7[-1:]
    x0=x0[["timeframe","datetime"]]
    x1=x1[["timeframe","datetime"]]
    x2=x2[["timeframe","datetime"]]
    x3=x3[["timeframe","datetime"]]
    x4=x4[["timeframe","datetime"]]
    x5=x5[["timeframe","datetime"]]
    x6=x6[["timeframe","datetime"]]
    new_df= pd.concat([x0,x1,x2,x3,x4,x5,x6])
    datetime=[]
    for dt in new_df["datetime"]:
        temp=dt.split(" ")
        st=temp[1].split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{} {}:{}:{}".format(temp[0],st[0],st[1],st[2])
        datetime.append(temp)
    new_df['datetime']=datetime
       
    tf_dataframe=pd.DataFrame(np.nan, index=[i for i in range(28)], columns=["pair","1mo","1wk","1d","60m","30m","15m","5m","1mo strength","1wk strength","1d strength","60m strength","30m strength","15m strength","5m strength"])
    tf_dataframe['5m']=new_df["datetime"].iloc[0]
    tf_dataframe['15m']=new_df["datetime"].iloc[1]
    tf_dataframe['30m']=new_df["datetime"].iloc[2]
    tf_dataframe['60m']=new_df["datetime"].iloc[3]
    tf_dataframe['1d']=new_df["datetime"].iloc[4]
    tf_dataframe['1wk']=new_df["datetime"].iloc[5]
    tf_dataframe['1mo']=new_df["datetime"].iloc[6]
      
    print(tf_dataframe)
    tf_dataframe.to_csv("strength_analysis.csv",index=False)
    

#strarting dash server
external_stylesheets = [dbc.themes.CERULEAN]
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)

#application layout
app.layout = html.Div(
    [
        #Navbar
        dbc.Navbar([
            html.A(
                dbc.Row([
                    dbc.Col(dbc.NavbarBrand("FOREX DASHBOARD", className="col-6",style={'font-size':'20px','fontColor': 'black'}),width={"size":1,"offset":5}),
                ],
                    style={},
                    )
                    
            )
        ],
        ),
        
        dbc.Row([
            #Dropdown strength
            dbc.Col(dcc.Dropdown(
                    id='strength-impulse-dropdown',
                    options=[
                        {'label': '5min', 'value': '5m' },
                        {'label': '15min', 'value': '15m'},
                        {'label': '30min', 'value': '30m'},
                        {'label': '60min', 'value': '60m'},
                        {'label': '1day', 'value': '1d'},
                        {'label': '1week', 'value': '1wk'},
                        {'label': '1month', 'value': '1mo'},

                    ],value='5m'),width={"size": 4},style={'fontColor': 'black'}
                ),
            dbc.Col(dash_table.DataTable(
                id='time_table',
                columns=[
                   {"name": 'Interval', "id": 'timeframe'},
                   {'name': 'Last Update', 'id': 'datetime'}
                ],
                page_current=0,
                page_size=7,
             ),width={"size": 5,'offset':3})
                
        ]),     
        #plotting graphs for strength and impulse
         dbc.Row([
                    dbc.Col(dcc.Graph(id='strength_graph'),width=6),
                    dbc.Col(dcc.Graph(id='impulse_graph'),width=6)
                ]),
        
        dbc.Row([
                    html.H4("Strength and Impulse Currency")
        ], justify="center", align="center", className="h-50"),
        #table for strength and impulse currency
        dbc.Row([
            dbc.Col(
                 dash_table.DataTable(
                    id='impulse_strength_currency',
                    columns=[{'name': 'pair', 'id': 'pair'},
                            {'name': '1mo strength', 'id': '1mo strength'},
                            {'name': '1wk strength', 'id': '1wk strength'},
                            {'name': '1d strength', 'id': '1d strength'},
                            {'name': '60m strength', 'id': '60m strength'},
                            {'name': '30m strength', 'id': '30m strength'},
                            {'name': '15m strength', 'id': '15m strength'},
                            {'name': '5m strength', 'id': '5m strength'},
                            {'name': '1mo impulse', 'id': '1mo impulse'},
                            {'name': '1wk impulse', 'id': '1wk impulse'},
                            {'name': '1d impulse', 'id': '1d impulse'},
                            {'name': '60m impulse', 'id': '60m impulse'},
                            {'name': '30m impulse', 'id': '30m impulse'},
                            {'name': '15m impulse', 'id': '15m impulse'},
                            {'name': '5m impulse', 'id': '5m impulse'}
                            ],
                    page_current=0,
                    page_size=8,
                    sort_action='custom',
                    sort_mode='multi',
                    filter_action='native',
                    sort_by=[]
                )
            )
        ]),


        dbc.Row([
                    html.H4("Strength Old v/s New Analysis")
                ], justify="center", align="center", className="h-50"),
        #table for strength and impulse currency
        dbc.Row([
            dbc.Col(
                 dash_table.DataTable(
                    id='strength_analysis',
                    columns=[#add
                                {'name': 'pair', 'id': 'pair'},
                                {'name': '1mo strength', 'id': '1mo'},
                                {'name': '1wk strength', 'id': '1wk'},
                                {'name': '1d strength', 'id': '1d'},
                                {'name': '60m strength', 'id': '60m'},
                                {'name': '30m strength', 'id': '30m'},
                                {'name': '15m strength', 'id': '15m'},
                                {'name': '5m strength', 'id': '5m'},
                            ],
                    page_current=0,
                    page_size=28,
                    sort_action='custom',
                    sort_mode='multi',
                    filter_action='native',
                    sort_by=[],
                    style_data_conditional=[
                                    ################
                                    {
                                        'if': {
                                            'column_id': '5m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{5m strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '5m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{5m strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    },
                                    ################
                                    {
                                        'if': {
                                            'column_id': '15m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{15m strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '15m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{15m strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    },
                                    ################
                                    {
                                        'if': {
                                            'column_id': '30m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{30m strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '30m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{30m strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    },
                                    ################
                                    {
                                        'if': {
                                            'column_id': '60m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{60m strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '60m',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{60m strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    },
                                    ################
                                    {
                                        'if': {
                                            'column_id': '1d',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{1d strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '1d',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{1d strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    },
                                    ################
                                    {
                                        'if': {
                                            'column_id': '1wk',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{1wk strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '1wk',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{1wk strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    },
                                    ################
                                    {
                                        'if': {
                                            'column_id': '1mo',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{1mo strength}} = {}'.format("red")
                                        },
                                        'backgroundColor': 'red',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {
                                            'column_id': '1mo',

                                            # since using .format, escape { with {{
                                            'filter_query': '{{1mo strength}} = {}'.format('green')
                                        },
                                        'backgroundColor': 'green',
                                        'color': 'white'
                                    }
                    ]

                )
            )
        ]),


        dbc.Row([
                    html.H4("Pair Buy Setup")
                ], justify="center", align="center", className="h-50"),
        #table for strength and impulse currency
        dbc.Row([
            dbc.Col(
                 dash_table.DataTable(
                    id='pair_buy_setup',
                    columns=[#add
                                {'name': 'pair', 'id': 'pair'},
                                {'name': '1mo strength', 'id': '1mo strength'},
                                {'name': '1wk strength', 'id': '1wk strength'},
                                {'name': '1d strength', 'id': '1d strength'},
                                {'name': '60m strength', 'id': '60m strength'},
                                {'name': '30m strength', 'id': '30m strength'},
                                {'name': '15m strength', 'id': '15m strength'},
                                {'name': '5m strength', 'id': '5m strength'},
                                {'name': '1mo impulse', 'id': '1mo impulse'},
                                {'name': '1wk impulse', 'id': '1wk impulse'},
                                {'name': '1d impulse', 'id': '1d impulse'},
                                {'name': '60m impulse', 'id': '60m impulse'},
                                {'name': '30m impulse', 'id': '30m impulse'},
                                {'name': '15m impulse', 'id': '15m impulse'},
                                {'name': '5m impulse', 'id': '5m impulse'}
                            ],
                    page_current=0,
                    page_size=28,
                    sort_action='custom',
                    filter_action='native',
                    sort_mode='multi',
                    sort_by=[]
                )
            )
        ]),

        dbc.Row([
                    html.H4("Pair Sell Setup")
                ], justify="center", align="center", className="h-50"),
        #table for strength and impulse currency
        dbc.Row([
            dbc.Col(
                 dash_table.DataTable(
                    id='pair_sell_setup',
                    columns=[#add
                                {'name': 'pair', 'id': 'pair'},
                                {'name': '1mo strength', 'id': '1mo strength'},
                                {'name': '1wk strength', 'id': '1wk strength'},
                                {'name': '1d strength', 'id': '1d strength'},
                                {'name': '60m strength', 'id': '60m strength'},
                                {'name': '30m strength', 'id': '30m strength'},
                                {'name': '15m strength', 'id': '15m strength'},
                                {'name': '5m strength', 'id': '5m strength'},
                                {'name': '1mo impulse', 'id': '1mo impulse'},
                                {'name': '1wk impulse', 'id': '1wk impulse'},
                                {'name': '1d impulse', 'id': '1d impulse'},
                                {'name': '60m impulse', 'id': '60m impulse'},
                                {'name': '30m impulse', 'id': '30m impulse'},
                                {'name': '15m impulse', 'id': '15m impulse'},
                                {'name': '5m impulse', 'id': '5m impulse'}
                            ],
                    page_current=0,
                    page_size=28,
                    sort_action='custom',
                    filter_action='native',
                    sort_mode='multi',
                    sort_by=[]
                )
            )
        ]),





        dbc.Row([
                    html.H4("Strength and Impulse Currency Pair")
                ], justify="center", align="center", className="h-50"),
        #table for strength and impulse currency
        dbc.Row([
            dbc.Col(
                 dash_table.DataTable(
                    id='impulse_strength_currency_pair',
                    columns=[#add
                                {'name': 'pair', 'id': 'pair'},
                                {'name': '1mo strength', 'id': '1mo strength'},
                                {'name': '1wk strength', 'id': '1wk strength'},
                                {'name': '1d strength', 'id': '1d strength'},
                                {'name': '60m strength', 'id': '60m strength'},
                                {'name': '30m strength', 'id': '30m strength'},
                                {'name': '15m strength', 'id': '15m strength'},
                                {'name': '5m strength', 'id': '5m strength'},
                                {'name': '1mo impulse', 'id': '1mo impulse'},
                                {'name': '1wk impulse', 'id': '1wk impulse'},
                                {'name': '1d impulse', 'id': '1d impulse'},
                                {'name': '60m impulse', 'id': '60m impulse'},
                                {'name': '30m impulse', 'id': '30m impulse'},
                                {'name': '15m impulse', 'id': '15m impulse'},
                                {'name': '5m impulse', 'id': '5m impulse'}
                            ],
                    page_current=0,
                    page_size=28,
                    sort_action='custom',
                    filter_action='native',
                    sort_mode='multi',
                    sort_by=[]
                )
            )
        ]),
        dcc.Interval(
            id='interval-component',
            interval=120*1000, # in milliseconds
            n_intervals=0
        )
              
    ]
)




#strength graph plotting
def min15data():
    #importing excel
    df1=pd.read_excel("15minfile_CS.xlsx")    
    datatime=[]
    temp=[]
    for dt in df1["datetime"]:        
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df1["time"]=datatime
    #plotting line graph
    df_infi= df1.iloc[-192:]
    df_infi1=df1.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'Strength'},
                        
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})

    return fig

def min5data():
    df2=pd.read_excel("5minfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df2["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df2["time"]=datatime
    df_infi= df2.iloc[-192:]
    df_infi1=df2.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'Strength'},
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def min30data():
    df3=pd.read_excel("30minfile_CS.xlsx")
    
    datatime=[]
    temp=[]
    for dt in df3["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df3["time"]=datatime
    df_infi= df3.iloc[-192:]
    df_infi1=df3.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'Strength'},
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def min60data():
    df4=pd.read_excel("60minfile_CS.xlsx")
    
    datatime=[]
    temp=[]
    for dt in df4["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df4["time"]=datatime
    df_infi= df4.iloc[-184:]
    df_infi1=df4.iloc[-192:-184].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'Strength'},
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def day1data():
    df5=pd.read_excel("1dayfile_CS.xlsx")
    
    datatime=[]
    temp=[]
    for dt in df5["datetime"]:
        temp=dt.split(" ")
        temp=temp[0]+"."
        # print(temp)
        datatime.append(temp)
    df5["time"]=datatime
    df_infi= df5.iloc[-192:]
    df_infi1=df5.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Dates"})
    fig.update_layout(yaxis={'title':'Strength'},
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def week1data():
    df6=pd.read_excel("1weekfile_CS.xlsx")
    
    datatime=[]
    temp=[]
    for dt in df6["datetime"]:
        temp=dt.split(" ")
        temp=temp[0]+"."
        # print(temp)
        datatime.append(temp)
    df6["time"]=datatime
    df_infi= df6.iloc[-192:]
    df_infi1=df6.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Dates"})
    fig.update_layout(yaxis={'title':'Strength'},
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def month1data():
    df7=pd.read_excel("1monthfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df7["datetime"]:
        temp=dt.split(" ")
        temp=temp[0]+"."
        # print(temp)
        datatime.append(temp)
    df7["time"]=datatime
    df_infi= df7.iloc[-192:]
    df_infi1=df7.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["strength"] ,color='currency' ,height=400,labels={"x": "Dates"})
    fig.update_layout(yaxis={'title':'Strength'},
                      title={'text':'Strength',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig
#impulse graph plotting

def min15dataimpulse():
    #importing excel
    df1=pd.read_excel("15minfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df1["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df1["time"]=datatime
    #plotting line graph
    df_infi= df1.iloc[-192:]
    df_infi1=df1.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})

    return fig

def min5dataimpulse():
    df2=pd.read_excel("5minfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df2["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df2["time"]=datatime
    df_infi= df2.iloc[-192:]
    df_infi1=df2.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def min30dataimpulse():
    df3=pd.read_excel("30minfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df3["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df3["time"]=datatime
    df_infi= df3.iloc[-192:]
    df_infi1=df3.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def min60dataimpulse():
    df4=pd.read_excel("60minfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df4["datetime"]:
        temp=dt.split(" ")
        temp=temp[1][:-3]
        st=temp.split(":")
        st[0]=(int(st[0])+2) % 24
        temp="{}:{}".format(st[0],st[1])
        # print(temp)
        datatime.append(temp)
    df4["time"]=datatime
    df_infi= df4.iloc[-184:]
    df_infi1=df4.iloc[-192:-184].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Time"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def day1dataimpulse():
    df5=pd.read_excel("1dayfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df5["datetime"]:
        temp=dt.split(" ")
        temp=temp[0]+"."
        # print(temp)
        datatime.append(temp)
    df5["time"]=datatime
    df_infi= df5.iloc[-192:]
    df_infi1=df5.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Dates"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def week1dataimpulse():
    df6=pd.read_excel("1weekfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df6["datetime"]:
        temp=dt.split(" ")
        temp=temp[0]+"."
        # print(temp)
        datatime.append(temp)
    df6["time"]=datatime
    df_infi= df6.iloc[-192:]
    df_infi1=df6.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Dates"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

def month1dataimpulse():
    df7=pd.read_excel("1monthfile_CS.xlsx")
    datatime=[]
    temp=[]
    for dt in df7["datetime"]:
        temp=dt.split(" ")
        temp=temp[0]+"."
        # print(temp)
        datatime.append(temp)
    df7["time"]=datatime
    df_infi= df7.iloc[-192:]
    df_infi1=df7.iloc[-200:-192].sort_values("currency")
    df_delta = pd.concat([df_infi1,df_infi],ignore_index=True)
    fig = px.line(df_delta, x=df_delta["time"], y=df_delta["ema_fast"] ,color='currency' ,height=400,labels={"x": "Dates"})
    fig.update_layout(yaxis={'title':'ema_fast'},
                      title={'text':'ema_fast(9)',
                      'font':{'size':20},'x':0.5,'xanchor':'center'})
    return fig

#callback for plotting strength graph
@app.callback(
    dash.dependencies.Output('strength_graph', 'figure'),
    [dash.dependencies.Input('strength-impulse-dropdown', 'value'),
    dash.dependencies.Input('interval-component', 'n_intervals')])
def update_output(value,n):
    #call function according to the value input
    if value =="15m":
        return min15data()
    if value =="5m":
        return min5data()
    if value =="30m":
        return min30data()
    if value=="60m":
        return min60data()
    if value=="1d":
        return day1data()
    if value == "1wk":
        return week1data()
    if value =="1mo":
        return month1data()

#callback for plotting impulse graph   
@app.callback(
    dash.dependencies.Output('impulse_graph', 'figure'),
    [dash.dependencies.Input('strength-impulse-dropdown', 'value'),
    dash.dependencies.Input('interval-component', 'n_intervals')])
def update_output(value,n):
    #call function according to the value input
    if value =="15m":
        return min15dataimpulse()
    if value =="5m":
        return min5dataimpulse()
    if value =="30m":
        return min30dataimpulse()
    if value=="60m":
        return min60dataimpulse()
    if value=="1d":
        return day1dataimpulse()
    if value == "1wk":
        return week1dataimpulse()
    if value =="1mo":
        return month1dataimpulse()

#showing table for strength and impulse currency
@app.callback(
    dash.dependencies.Output('impulse_strength_currency', "data"),
    dash.dependencies.Output('impulse_strength_currency','style_data_conditional'),
    dash.dependencies.Input('impulse_strength_currency', "page_current"),
    dash.dependencies.Input('impulse_strength_currency', "page_size"),
    dash.dependencies.Input('impulse_strength_currency', "sort_by"),
    dash.dependencies.Input('strength-impulse-dropdown','value'),
    dash.dependencies.Input('interval-component', 'n_intervals'))
def update_table(page_current, page_size, sort_by,value1,n):
    #create dataframe
    df=pd.DataFrame()
    df=show_currency()
    #adding styles for strength and impulse
    (styles1, legend) = discrete_background_color_bins_reds(df,columns=["1mo strength"])
    (styles2, legend) = discrete_background_color_bins_reds(df,columns=['1wk strength'])
    (styles3, legend) = discrete_background_color_bins_reds(df,columns=["15m strength"])
    (styles4, legend) = discrete_background_color_bins_reds(df,columns=['5m strength'])
    (styles5, legend) = discrete_background_color_bins_reds(df,columns=['30m strength'])
    (styles6, legend) = discrete_background_color_bins_reds(df,columns=['1d strength'])
    (styles7, legend) = discrete_background_color_bins_reds(df,columns=['60m strength'])

    (styles8, legend) = discrete_background_color_bins_blues(df,columns=["1mo impulse"])
    (styles9, legend) = discrete_background_color_bins_blues(df,columns=['1wk impulse'])
    (styles10, legend) = discrete_background_color_bins_blues(df,columns=["15m impulse"])
    (styles11, legend) = discrete_background_color_bins_blues(df,columns=['5m impulse'])
    (styles12, legend) = discrete_background_color_bins_blues(df,columns=['30m impulse'])
    (styles13, legend) = discrete_background_color_bins_blues(df,columns=['1d impulse'])
    (styles14, legend) = discrete_background_color_bins_blues(df,columns=['60m impulse'])
    #combine all styles
    style=styles1+styles2+styles3+styles4+styles5+styles6+styles7+styles8+styles9+styles10+styles11+styles12+styles13+styles14
    print(sort_by)
    if len(sort_by):
        dff = df.sort_values(
            [col['column_id'] for col in sort_by],
            ascending=[
                col['direction'] == 'asc'
                for col in sort_by
            ],
            inplace=False
        )
    else:
        # No sort is applied
        dff = df
    #return table as well as style
    return [dff.iloc[
        page_current*page_size:(page_current+ 1)*page_size
    ].to_dict('records'),style]

#showing table for strength and impulse currency pair
@app.callback(
    dash.dependencies.Output('impulse_strength_currency_pair', "data"),
    dash.dependencies.Output('impulse_strength_currency_pair','style_data_conditional'),
    dash.dependencies.Input('impulse_strength_currency_pair', "page_current"),
    dash.dependencies.Input('impulse_strength_currency_pair', "page_size"),
    dash.dependencies.Input('impulse_strength_currency_pair', "sort_by"),
    dash.dependencies.Input('strength-impulse-dropdown','value'),
    dash.dependencies.Input('interval-component', 'n_intervals'))
def update_table(page_current, page_size, sort_by,value1,n):
    #create dataframe
    df=pd.DataFrame()
    df=show_currency_pair()
    #adding styles for strength and impulse
    (styles1, legend) = discrete_background_color_bins_reds(df,columns=["1mo strength"])
    (styles2, legend) = discrete_background_color_bins_reds(df,columns=['1wk strength'])
    (styles3, legend) = discrete_background_color_bins_reds(df,columns=["15m strength"])
    (styles4, legend) = discrete_background_color_bins_reds(df,columns=['5m strength'])
    (styles5, legend) = discrete_background_color_bins_reds(df,columns=['30m strength'])
    (styles6, legend) = discrete_background_color_bins_reds(df,columns=['1d strength'])
    (styles7, legend) = discrete_background_color_bins_reds(df,columns=['60m strength'])

    (styles8, legend) = discrete_background_color_bins_blues(df,columns=["1mo impulse"])
    (styles9, legend) = discrete_background_color_bins_blues(df,columns=['1wk impulse'])
    (styles10, legend) = discrete_background_color_bins_blues(df,columns=["15m impulse"])
    (styles11, legend) = discrete_background_color_bins_blues(df,columns=['5m impulse'])
    (styles12, legend) = discrete_background_color_bins_blues(df,columns=['30m impulse'])
    (styles13, legend) = discrete_background_color_bins_blues(df,columns=['1d impulse'])
    (styles14, legend) = discrete_background_color_bins_blues(df,columns=['60m impulse'])
    #combine all styles
    style=styles1+styles2+styles3+styles4+styles5+styles6+styles7+styles8+styles9+styles10+styles11+styles12+styles13+styles14
    print(sort_by)
    if len(sort_by):
        dff = df.sort_values(
            [col['column_id'] for col in sort_by],
            ascending=[
                col['direction'] == 'asc'
                for col in sort_by
            ],
            inplace=False
        )
    else:
        # No sort is applied
        dff = df
    #return table as well as style
    return [dff.iloc[
        page_current*page_size:(page_current+ 1)*page_size
    ].to_dict('records'),style]

@app.callback(
    dash.dependencies.Output('time_table', "data"),
    dash.dependencies.Input('time_table', "page_current"),
    dash.dependencies.Input('time_table', "page_size"),
    dash.dependencies.Input('interval-component', 'n_intervals'),
    dash.dependencies.Input('strength-impulse-dropdown','value'),
)
def time_table(page_current, page_size,n,v):
    new_df=last_update()
    return new_df.iloc[page_current*page_size:(page_current+ 1)*page_size].to_dict('records')


@app.callback(
    dash.dependencies.Output('pair_buy_setup', "data"),
    dash.dependencies.Output('pair_buy_setup','style_data_conditional'),
    dash.dependencies.Input('pair_buy_setup', "page_current"),
    dash.dependencies.Input('pair_buy_setup', "page_size"),
    dash.dependencies.Input('pair_buy_setup', "sort_by"),
    dash.dependencies.Input('strength-impulse-dropdown','value'),
    dash.dependencies.Input('interval-component', 'n_intervals')
)
def pair_buy_setup(page_current, page_size, sort_by,value1,n):
    df=pd.DataFrame()
    df=show_currency_pair()
    df=df[df['1wk strength']>0]
    #adding styles for strength and impulse
    (styles1, legend) = discrete_background_color_bins_reds(df,columns=["1mo strength"])
    (styles2, legend) = discrete_background_color_bins_reds(df,columns=['1wk strength'])
    (styles3, legend) = discrete_background_color_bins_reds(df,columns=["15m strength"])
    (styles4, legend) = discrete_background_color_bins_reds(df,columns=['5m strength'])
    (styles5, legend) = discrete_background_color_bins_reds(df,columns=['30m strength'])
    (styles6, legend) = discrete_background_color_bins_reds(df,columns=['1d strength'])
    (styles7, legend) = discrete_background_color_bins_reds(df,columns=['60m strength'])

    (styles8, legend) = discrete_background_color_bins_blues(df,columns=["1mo impulse"])
    (styles9, legend) = discrete_background_color_bins_blues(df,columns=['1wk impulse'])
    (styles10, legend) = discrete_background_color_bins_blues(df,columns=["15m impulse"])
    (styles11, legend) = discrete_background_color_bins_blues(df,columns=['5m impulse'])
    (styles12, legend) = discrete_background_color_bins_blues(df,columns=['30m impulse'])
    (styles13, legend) = discrete_background_color_bins_blues(df,columns=['1d impulse'])
    (styles14, legend) = discrete_background_color_bins_blues(df,columns=['60m impulse'])
    #combine all styles
    style=styles1+styles2+styles3+styles4+styles5+styles6+styles7+styles8+styles9+styles10+styles11+styles12+styles13+styles14
    print(sort_by)
    if len(sort_by):
        dff = df.sort_values(
            [col['column_id'] for col in sort_by],
            ascending=[
                col['direction'] == 'asc'
                for col in sort_by
            ],
            inplace=False
        )
    else:
        # No sort is applied
        dff = df
    #return table as well as style
    return [dff.iloc[
        page_current*page_size:(page_current+ 1)*page_size
    ].to_dict('records'),style]
    

@app.callback(
    dash.dependencies.Output('pair_sell_setup', "data"),
    dash.dependencies.Output('pair_sell_setup','style_data_conditional'),
    dash.dependencies.Input('pair_sell_setup', "page_current"),
    dash.dependencies.Input('pair_sell_setup', "page_size"),
    dash.dependencies.Input('pair_sell_setup', "sort_by"),
    dash.dependencies.Input('strength-impulse-dropdown','value'),
    dash.dependencies.Input('interval-component', 'n_intervals')
)
def pair_sell_setup(page_current, page_size, sort_by,value1,n):
    df=pd.DataFrame()
    df=show_currency_pair()
    df=df[df['1wk strength']<0]
    #adding styles for strength and impulse
    (styles1, legend) = discrete_background_color_bins_reds(df,columns=["1mo strength"])
    (styles2, legend) = discrete_background_color_bins_reds(df,columns=['1wk strength'])
    (styles3, legend) = discrete_background_color_bins_reds(df,columns=["15m strength"])
    (styles4, legend) = discrete_background_color_bins_reds(df,columns=['5m strength'])
    (styles5, legend) = discrete_background_color_bins_reds(df,columns=['30m strength'])
    (styles6, legend) = discrete_background_color_bins_reds(df,columns=['1d strength'])
    (styles7, legend) = discrete_background_color_bins_reds(df,columns=['60m strength'])

    (styles8, legend) = discrete_background_color_bins_blues(df,columns=["1mo impulse"])
    (styles9, legend) = discrete_background_color_bins_blues(df,columns=['1wk impulse'])
    (styles10, legend) = discrete_background_color_bins_blues(df,columns=["15m impulse"])
    (styles11, legend) = discrete_background_color_bins_blues(df,columns=['5m impulse'])
    (styles12, legend) = discrete_background_color_bins_blues(df,columns=['30m impulse'])
    (styles13, legend) = discrete_background_color_bins_blues(df,columns=['1d impulse'])
    (styles14, legend) = discrete_background_color_bins_blues(df,columns=['60m impulse'])
    #combine all styles
    style=styles1+styles2+styles3+styles4+styles5+styles6+styles7+styles8+styles9+styles10+styles11+styles12+styles13+styles14
    print(sort_by)
    if len(sort_by):
        dff = df.sort_values(
            [col['column_id'] for col in sort_by],
            ascending=[
                col['direction'] == 'asc'
                for col in sort_by
            ],
            inplace=False
        )
    else:
        # No sort is applied
        dff = df
    #return table as well as style
    return [dff.iloc[
        page_current*page_size:(page_current+ 1)*page_size
    ].to_dict('records'),style]
    
@app.callback(
    dash.dependencies.Output('strength_analysis', "data"),
    dash.dependencies.Input('strength_analysis', "page_current"),
    dash.dependencies.Input('strength_analysis', "page_size"),
    dash.dependencies.Input('strength_analysis', "sort_by"),
    dash.dependencies.Input('strength-impulse-dropdown','value'),
    dash.dependencies.Input('interval-component', 'n_intervals')    
)
def strength_analysis(page_current, page_size,sort_by,value1,n):
    last_updates=last_update()

    old_df= pd.read_csv('old_data.csv')
    new_df= pd.read_csv('new_data.csv')
    
    tf_dataframe=pd.read_csv('strength_analysis.csv')
    tf_dataframe['pair']=old_df['pair']
    tf_dataframe['5m strength'] = np.where((old_df['5m strength'] > 0) & ( new_df['5m strength'] >0), "green",np.where((old_df['5m strength'] < 0) & ( new_df['5m strength'] >0), "green", np.where((old_df['5m strength'] > 0) & ( new_df['5m strength'] <0), "red", "red" )) )
    tf_dataframe['15m strength'] = np.where((old_df['15m strength'] > 0) & ( new_df['15m strength'] >0), "green",np.where((old_df['15m strength'] < 0) & ( new_df['15m strength'] >0), "green", np.where((old_df['15m strength'] > 0) & ( new_df['15m strength'] <0), "red", "red" )) )
    tf_dataframe['30m strength'] =np.where((old_df['30m strength'] > 0) & ( new_df['30m strength'] >0), "green",np.where((old_df['30m strength'] < 0) & ( new_df['30m strength'] >0), "green", np.where((old_df['30m strength'] > 0) & ( new_df['30m strength'] <0), "red", "red" )) )
    tf_dataframe['60m strength'] = np.where((old_df['60m strength'] > 0) & ( new_df['60m strength'] >0), "green",np.where((old_df['60m strength'] < 0) & ( new_df['60m strength'] >0), "green", np.where((old_df['60m strength'] > 0) & ( new_df['60m strength'] <0), "red", "red" )) )
    tf_dataframe['1d strength'] = np.where((old_df['1d strength'] > 0) & ( new_df['1d strength'] >0), "green",np.where((old_df['1d strength'] < 0) & ( new_df['1d strength'] >0), "green", np.where((old_df['1d strength'] > 0) & ( new_df['1d strength'] <0), "red", "red" )) )
    tf_dataframe['1mo strength'] = np.where((old_df['1mo strength'] > 0) & ( new_df['1mo strength'] >0), "green",np.where((old_df['1mo strength'] < 0) & ( new_df['1mo strength'] >0), "green", np.where((old_df['1mo strength'] > 0) & ( new_df['1mo strength'] <0), "red", "red" )) )
    tf_dataframe['1wk strength'] = np.where((old_df['1wk strength'] > 0) & ( new_df['1wk strength'] >0), "green",np.where((old_df['1wk strength'] < 0) & ( new_df['1wk strength'] >0), "green", np.where((old_df['1wk strength'] > 0) & ( new_df['1wk strength'] <0), "red", "red" )) )
    data=[]
    column_list = list(tf_dataframe)
    for i in column_list[1:8]:   
        files="{} strength".format(i)     
        if i=="5m":
            k=0
        if i=="15m":
            k=1
        if i=="30m":
            k=2
        if i=="60m":
            k=3
        if i=="1d":
            k=4
        if i=="1wk":
            k=5
        if i=="1mo":
            k=6
        for j in range(28):        
            if old_df[files].iloc[j] >0 and new_df[files].iloc[j]<0:
                tf_dataframe[i][j] = last_updates["datetime"].iloc[k]
                x="{} {} {}".format(tf_dataframe[i][j],i,j)
                data.append(x)
            if old_df[files].iloc[j] <0 and new_df[files].iloc[j]>0:
                tf_dataframe[i][j] = last_updates["datetime"].iloc[k]
                x="{} {} {}".format(tf_dataframe[i][j],i,j)
                data.append(x)
    #print changed data
    print("+++++++++++++++++++++++++++++++++++++++++++++++++")
    print(data)
    print("+++++++++++++++++++++++++++++++++++++++++++++++++")
    print(sort_by)


    tf_dataframe.to_csv("strength_analysis.csv",index=False)
    # for col in sort_by:

    if len(sort_by):
        if len(sort_by):
            for col in sort_by:
                tf_dataframe[col['column_id']] = pd.to_datetime(tf_dataframe[col['column_id']], dayfirst=True)

        dff = tf_dataframe.sort_values(
            [col['column_id'] for col in sort_by],
            ascending=[
                col['direction'] == 'asc'
                for col in sort_by
            ],
            inplace=False
        )
        for col in sort_by:
            dff[col['column_id']] = dff[col['column_id']].dt.strftime('%Y-%m-%d %H:%M:%S')
    else:
        # No sort is applied
        dff = tf_dataframe

    return dff.iloc[
        page_current*page_size:(page_current+ 1)*page_size
    ].to_dict('records')




    

if __name__ == '__main__':
    
    # #generating historical data
    h1 = multiprocessing.Process(target=historical_data_generator, args=["5m","5minfile",0,29])
    h2 = multiprocessing.Process(target=historical_data_generator, args=["15m","15minfile",0,29])
    h3 = multiprocessing.Process(target=historical_data_generator, args=["30m","30minfile",0,29])
    h4 = multiprocessing.Process(target=historical_data_generator, args=["60m","60minfile",0,29])
    h5 = multiprocessing.Process(target=historical_data_generator, args=["1d","1dayfile",0,29])
    h6 = multiprocessing.Process(target=historical_data_generator, args=["1wk","1weekfile",0,29])
    h7 = multiprocessing.Process(target=historical_data_generator, args=["1mo","1monthfile",0,29])
    print("Start with Generating Historical Data")
    h1.start()
    h2.start()
    h3.start()
    h4.start()
    h5.start()
    h6.start()
    h7.start()
    
    h1.join()
    print("5m file generated")
    h2.join()
    print("15m file generated")
    h3.join()
    print("30m file generated")
    h4.join()
    print("60m file generated")
    h5.join()
    print("1d file generated")
    h6.join()
    print("1wk file generated")
    h7.join()
    print("1mo file generated")

    #create two csv file for old_data and new_data
    temp = pd.DataFrame()
    temp.to_csv("old_data.csv",index=False)
    temp.to_csv("new_data.csv",index=False)
    
    temp = show_currency_pair()
    df= temp[['pair','1mo strength','1wk strength','1d strength','60m strength','30m strength','15m strength','5m strength']]
    df.to_csv('new_data.csv')
    df.to_csv('old_data.csv')
    old_update()

    

    
    
    
    #running code for each timeframe
    #creating multiprocesses for each timeframe

    p1 = multiprocessing.Process(target=runProgram, args=['5m',"5minfile",300,10000])
    p2 = multiprocessing.Process(target=runProgram, args=['15m',"15minfile",900,10000])
    p3 = multiprocessing.Process(target=runProgram, args=['30m',"30minfile",1800,10000])
    p4 = multiprocessing.Process(target=runProgram, args=['60m',"60minfile",3600,10000])
    p5 = multiprocessing.Process(target=runProgram, args=['1d',"1dayfile",86400,10000])
    p6 = multiprocessing.Process(target=runProgram, args=['1wk',"1weekfile",604800,10000])
    p7 = multiprocessing.Process(target=runProgram, args=['1mo',"1monthfile",(864000*3),10000])

    p1.start() # starting workers
    p2.start() # starting workers
    p3.start() # starting workers
    p4.start() # starting workers
    p5.start() # starting workers
    p6.start() # starting workers
    p7.start() # starting workers

    
    # #start dash server
    app.run_server(port=6262)
